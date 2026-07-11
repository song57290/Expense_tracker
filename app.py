from flask import Flask, render_template, request, redirect, url_for, send_from_directory, send_file, jsonify, abort, session
from functools import wraps
from models import db, Transaction, Budget, Category, Card, User, Savings, Investment, Notice, HelpItem, AppConfig, SalaryConfig, BudgetAllocation, FixedExpense, SavingsDeposit
from datetime import datetime, timedelta, timezone
from collections import defaultdict
import openpyxl
import xlrd
from io import BytesIO
import tempfile, json, os, re, base64, random, smtplib
from email.mime.text import MIMEText

_listing_cache = {}
_listing_ts = {}

def _get_stock_listing(market):
    import time, FinanceDataReader as fdr
    now = time.time()
    if market not in _listing_cache or now - _listing_ts.get(market, 0) > 86400:
        _listing_cache[market] = fdr.StockListing(market)
        _listing_ts[market] = now
    return _listing_cache[market]

app = Flask(__name__)
_DATA_DIR_ENV = os.environ.get('DATA_DIR', '')
if _DATA_DIR_ENV:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(_DATA_DIR_ENV, 'expense.db')
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///expense.db'
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-key-change-me-in-prod-2026x')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
db.init_app(app)

# ── Auth helpers ──────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated

def _seed_user_categories(uid):
    if Category.query.filter_by(user_id=uid).first():
        return
    defaults_expense = [('식사','🍚'),('간식','🍪'),('쇼핑','🛍️'),('자동차','🚗'),('교통','🚌'),('의료','💊'),('기타','📦')]
    defaults_income = [('급여','💰'),('부업','💼'),('용돈','🎁'),('이자','🏦'),('기타수입','📥')]
    for i, (name, icon) in enumerate(defaults_expense):
        db.session.add(Category(name=name, icon=icon, position=i, cat_type='expense', user_id=uid))
    for i, (name, icon) in enumerate(defaults_income):
        db.session.add(Category(name=name, icon=icon, position=len(defaults_expense)+i, cat_type='income', user_id=uid))
    db.session.commit()

# ── Template filters ──────────────────────────────────────────────────────────

@app.template_filter('bank_color')
def bank_color_filter(card_name):
    if not card_name:
        return 'background:#6c757d;color:white;'
    mappings = [
        ('신한', '#0046A0', 'white'), ('KB', '#FFB800', '#333'), ('국민', '#FFB800', '#333'),
        ('농협', '#009900', 'white'), ('NH', '#009900', 'white'), ('하나', '#009A8C', 'white'),
        ('우리', '#0069C8', 'white'), ('기업', '#005BB5', 'white'), ('IBK', '#005BB5', 'white'),
        ('카카오', '#FAE100', '#333'), ('토스', '#0064FF', 'white'), ('케이뱅크', '#00B4B4', 'white'),
        ('K뱅크', '#00B4B4', 'white'), ('SC', '#1B5DA0', 'white'), ('제일', '#1B5DA0', 'white'),
        ('씨티', '#003087', 'white'), ('iM', '#E8182C', 'white'), ('IM', '#E8182C', 'white'),
        ('수협', '#009ABF', 'white'), ('KDB', '#003087', 'white'), ('산업', '#003087', 'white'),
        ('BNK', '#0057A8', 'white'), ('부산', '#0057A8', 'white'), ('우체국', '#D40511', 'white'),
        ('SBI', '#E8391D', 'white'), ('신협', '#005BAB', 'white'), ('BC', '#D60B2F', 'white'),
        ('현대', '#1A1A1A', 'white'), ('롯데', '#CC0000', 'white'), ('삼성', '#005BAB', 'white'),
    ]
    for keyword, bg, fg in mappings:
        if keyword in card_name:
            return f'background:{bg};color:{fg};'
    return 'background:#6c757d;color:white;'

@app.template_filter('bank_logo')
def bank_logo_filter(card_name):
    mappings = [
        ('신한', '/static/cards/sinhanbank.png'), ('KB', '/static/cards/kbbank.png'),
        ('국민', '/static/cards/kbbank.png'), ('농협', '/static/cards/nhbank.png'),
        ('NH', '/static/cards/nhbank.png'), ('하나', '/static/cards/hanabank.png'),
        ('우리', '/static/cards/wooribank.png'), ('기업', '/static/cards/ibkbank.png'),
        ('IBK', '/static/cards/ibkbank.png'), ('카카오', '/static/cards/kakaobank.png'),
        ('토스', '/static/cards/tossbank.png'), ('케이뱅크', '/static/cards/kbank.png'),
        ('K뱅크', '/static/cards/kbank.png'), ('SC', '/static/cards/scbank.png'),
        ('제일', '/static/cards/scbank.png'), ('씨티', '/static/cards/citibank.png'),
        ('citi', '/static/cards/citibank.png'), ('IM', '/static/cards/imbank.png'),
        ('iM', '/static/cards/imbank.png'), ('수협', '/static/cards/suhyupbank.png'),
        ('KDB', '/static/cards/kdbbank.png'), ('산업', '/static/cards/kdbbank.png'),
        ('BNK', '/static/cards/bnkbank.png'), ('부산', '/static/cards/bnkbank.png'),
        ('우체국', '/static/cards/epostbank.png'), ('SBI', '/static/cards/sbibank.png'),
        ('신협', '/static/cards/cubank.png'), ('BC', '/static/banks/bccard.png'),
        ('현대', '/static/banks/hyundaicard.png'), ('롯데', '/static/banks/lottecard.png'),
        ('삼성', '/static/banks/samsungcard.png'),
    ]
    for keyword, path in mappings:
        if keyword in card_name:
            return path
    return None

# ── DB init ───────────────────────────────────────────────────────────────────

with app.app_context():
    db.create_all()
    from sqlalchemy import text
    for col, default in [('tier1', 20), ('tier2', 50), ('tier3', 80)]:
        try:
            with db.engine.connect() as conn:
                conn.execute(text(f"ALTER TABLE card ADD COLUMN {col} INTEGER DEFAULT {default}"))
                conn.commit()
        except Exception:
            pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE savings ADD COLUMN interest_type VARCHAR(10) NOT NULL DEFAULT '단리'"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE savings ADD COLUMN tax_type VARCHAR(10) NOT NULL DEFAULT '일반과세'"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE savings ADD COLUMN notify_day INTEGER"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE savings ADD COLUMN auto_tx BOOLEAN DEFAULT 0"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE savings ADD COLUMN auto_tx_day INTEGER"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE savings ADD COLUMN auto_tx_card VARCHAR(50) DEFAULT ''"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE savings ADD COLUMN manual_count INTEGER"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE savings ADD COLUMN is_paused BOOLEAN NOT NULL DEFAULT 0"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE savings ADD COLUMN bonus_amount INTEGER"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE investment ADD COLUMN exchange_rate FLOAT"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE category ADD COLUMN position INTEGER DEFAULT 0"))
            conn.commit()
        cats = Category.query.order_by(Category.id).all()
        for i, c in enumerate(cats):
            c.position = i
        db.session.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE category ADD COLUMN cat_type VARCHAR(10) DEFAULT 'expense'"))
            conn.commit()
    except Exception:
        pass
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE card ADD COLUMN account_balance INTEGER DEFAULT 0"))
            conn.commit()
    except Exception:
        pass
    # user_id column migration ("transaction" must be quoted — SQLite reserved word)
    for table_name in ['"transaction"', 'card', 'category', 'budget']:
        try:
            with db.engine.connect() as conn:
                conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN user_id INTEGER DEFAULT 1"))
                conn.commit()
        except Exception:
            pass
    for col in ['reset_code VARCHAR(6)', 'reset_expires DATETIME', 'nickname VARCHAR(50)']:
        try:
            with db.engine.connect() as conn:
                conn.execute(text(f"ALTER TABLE user ADD COLUMN {col}"))
                conn.commit()
        except Exception:
            pass
    for _sql in [
        "ALTER TABLE fixed_expense ADD COLUMN auto_register BOOLEAN DEFAULT 0",
        "ALTER TABLE fixed_expense ADD COLUMN tx_type VARCHAR(20) DEFAULT 'expense'",
        "ALTER TABLE fixed_expense ADD COLUMN tx_card VARCHAR(50) DEFAULT ''",
        'ALTER TABLE "transaction" ADD COLUMN exclude_perf BOOLEAN NOT NULL DEFAULT 0',
        "ALTER TABLE category ADD COLUMN exclude_perf BOOLEAN NOT NULL DEFAULT 0",
        "ALTER TABLE investment ADD COLUMN account_type VARCHAR(20) NOT NULL DEFAULT '일반'",
        "ALTER TABLE card ADD COLUMN linked_account_id INTEGER",
        'ALTER TABLE "transaction" ADD COLUMN exclude_stats BOOLEAN NOT NULL DEFAULT 0',
        "ALTER TABLE category ADD COLUMN exclude_stats BOOLEAN NOT NULL DEFAULT 0",
    ]:
        try:
            with db.engine.connect() as conn:
                conn.execute(text(_sql)); conn.commit()
        except Exception:
            pass
    db.create_all()
    # seed categories for user 1 (existing data owner)
    _seed_user_categories(1)

    # seed / upgrade help items
    _help_version = 'ver2.29'
    _help_defaults = [
        ('🏠', '홈', '수입·지출 내역을 기록하고 이번 달 내역을 관리합니다.\n\n• 이번 달 내역만 목록에 표시됩니다\n• 내역 목록은 날짜별로 그룹화되어 표시\n• 항목을 오른쪽으로 스와이프 → 수정 (PC에서는 마우스 드래그)\n• 항목을 왼쪽으로 스와이프 → 삭제 (PC에서는 마우스 드래그)\n• 카드/계좌를 지정하면 예산 탭 잔고에 자동 반영\n\n💱 카드 실적 제외\n• 내역 추가·수정 시 지출 항목에 "카드 실적에서 제외" 토글 제공\n• 계좌이체·대출 상환 등 실적에 포함되지 않는 거래에 사용\n• 제외 설정된 거래는 목록에 "실적제외" 배지로 표시\n• 카테고리에서 실적 제외 설정 시 해당 카테고리 내역 추가 시 자동으로 토글이 켜짐\n\n📥 내역 가져오기\n• 문자 붙여넣기: 카드·은행 문자를 붙여넣으면 자동 인식\n• 엑셀 업로드: 양식 다운로드 후 작성하거나 은행 내보내기 파일 바로 업로드\n  - 업로드 후 카테고리 및 카드/계좌 선택 화면으로 이동\n  - 일괄 적용: 지출/수입/카드 전체에 한 번에 지정 가능\n  - 카드 미선택 시 현금/미지정으로 저장\n  - 현금을 별도 추적하려면 예산 탭에서 현금 자산을 먼저 등록\n  - 오류 항목은 별도 표시 → 내용 확인 후 직접 수동 입력'),
        ('💳', '예산', '카드·은행·현금 잔고와 예적금·투자를 한눈에 확인합니다.\n\n• 자산 추가: 카드/은행 또는 현금 선택 후 등록\n• 초기 잔고: 앱 사용 시작 전 보유 금액 입력\n• 💸 대출·빚 등록: 초기 잔고에 음수(-) 입력 또는 "대출/빚" 유형 선택 → 잔고가 음수로 표시됨\n• 오른쪽 스와이프 → 수정, 왼쪽 스와이프 → 삭제 (PC에서는 마우스 드래그)\n• 수정 시 색상 구간 설정 가능: 빨강 ≤ / 노랑 ≤ / 파랑 ≤ / 초록 기준을 % 단위로 직접 조정\n\n🔗 연결 계좌 (카드 공유 잔고)\n• 하나의 은행 계좌에 여러 카드를 연결 가능\n• 연결된 카드는 잔고를 계좌와 공유하며, 실적·목표는 카드별로 분리 관리\n• 자산 추가 → "💳 카드/은행" → "연결 계좌" 드롭다운에서 연결할 계좌 선택\n  - 선택 안 하면 독립 계좌로 등록됨\n• 연결된 카드는 해당 계좌 아래 들여쓰기로 표시되며 🔗 배지로 구분\n\n🏦 예금·적금·청약\n• 만기일·이율·납입액 입력 시 이자 자동 계산 (단리/복리·세금 종류 선택)\n• 세금 종류: 일반과세(15.4%) / 세금우대(9.9%) / ISA / 비과세\n  - ISA: 신탁형 ISA 내 예금·적금에 적용 (중개형·일임형은 예금·적금 불가)\n  - ISA 일반형: 비과세 한도 200만원, 초과분 9.9% 분리과세\n  - ISA 서민형: 비과세 한도 400만원, 초과분 9.9% 분리과세\n• 청약도 연 이율·단리/복리·세금 종류 설정 가능 (비과세 기본)\n• 자동이체 등록: 이체일 설정 후 해당 날짜에 앱 열면 확인 팝업 → 등록 시 거래 자동 기록\n• 청약 납입일 알림: 납입일(몇 일) 입력 시 매월 해당 날짜 오전 9시에 푸시 알림 자동 발송 (알림 ON 필요)\n• 청약 추가 입금: 청약 카드에서 추가 입금 내역 기록, 잔고에 자동 합산\n• ✏ 납입 회차 수동 설정: 청약 카드에서 현재 납입 회차를 직접 수정 가능 (수동 설정 시 보라색 "수동" 표시)\n• ⏸ 일시정지: 청약·적금 카드에서 납입 일시정지 — 일시정지 중에는 자동이체 팝업과 고정 지출에서 자동 제외\n• 🎁 정부 지원금: 청년도약계좌 등 월 정부 지원금 등록 시 만기 수령액에 자동 합산 (비과세 처리)\n\n📈 투자\n• 종목·수량·매수가 입력, 티커로 현재가 자동 조회\n• 해외주식은 달러($) 기준으로 평단가·현재가 표시\n• 계좌 종류 선택: 일반 / ISA / 연금저축 / IRP → 종목 카드에 배지로 표시'),
        ('📅', '캘린더', '날짜별 수입·지출을 달력으로 확인합니다.\n\n• 날짜에 보라색 점(지출) / 초록 점(수입) 표시\n• 날짜 클릭 → 해당일 내역 팝업\n• 상단 년/월 클릭 → 원하는 달로 이동\n\n📋 월별 내역 목록\n• 달력 하단에 해당 월 전체 내역이 날짜별로 그룹화되어 표시\n• 전체 / 지출 / 수입 탭으로 필터링 가능\n• 달 이동 시 목록 자동 갱신'),
        ('📊', '통계', '카테고리별 지출과 자산 흐름을 차트로 분석합니다.\n\n• 도넛 차트: 카테고리별 지출 비중 시각화\n• 월별 추이: 지출 / 수입 / 전체 모드 전환 가능\n  - 전체 모드: 수입·지출 막대를 나란히 비교\n  - 날짜 범위 자유 설정, 전체 버튼으로 첫 거래부터 현재까지 한 번에 확인\n• 자산 구성: 현금·예금·적금/청약·투자를 도넛 차트로 표시\n  - 중앙 숫자는 대출 차감 후 순자산\n  - 대출 계좌가 있으면 목록 하단에 빨간색으로 별도 표시\n• 총 자산 추이: 전월 대비 어느 자산이 얼마나 변화했는지 항목별로 확인\n  - 차트 탭 하여 자세히 보기 → 월별 변화액을 클릭하면 카테고리별 세부 변화 표시\n• 월 이동 버튼으로 과거 달 조회 가능\n• 통계에서 제외한 거래·카테고리는 모든 집계에서 자동 제외'),
        ('💰', '월급', '월급 기준으로 예산을 계획하고 고정 지출을 관리합니다.\n\n💵 월급 설정\n• 월급 금액·지급일 입력 후 저장\n\n📋 예산 배분\n• "+ 카테고리 추가" 버튼으로 예산 잡을 카테고리만 선택해서 추가\n• 원화(₩)로 직접 입력, 월급 대비 % 자동 표시\n• 합계가 월급을 초과하면 경고 표시\n• 이번 달 실제 지출과 배분 예산을 나란히 비교\n• ⠿ 핸들을 드래그하여 카테고리 순서 변경 가능\n\n🔒 고정 지출\n• 매달 반복되는 지출 항목 등록 (구독, 보험, 관리비 등)\n• 자동 등록 ON: 지정일에 앱 열면 확인 팝업 → 등록 시 거래 자동 기록\n• 자동이체 설정한 적금·청약도 고정 지출 탭에 자동 표시'),
        ('🏷️', '카테고리', '지출·수입 카테고리를 관리합니다.\n(설정 탭 → 카테고리 관리 버튼으로 접근)\n\n• 추가 양식: 이모지·이름·지출수입·저장·취소를 한 줄로 입력\n• 이모지는 선택 사항 (비워도 저장 가능)\n• 왼쪽 핸들(⠿)을 드래그하여 순서 변경 (모바일·PC 모두 지원)\n• 드래그 중에는 수정/삭제 패널이 숨겨집니다\n• 수정 중인 항목은 앞으로 나오는 강조 효과로 표시\n• 항목을 오른쪽으로 스와이프 → 이름/이모지 수정\n• 항목을 왼쪽으로 스와이프 → 삭제\n• 지출/수입 탭 분리 관리\n\n💱 카드 실적 제외\n• 지출 카테고리에 "카드 실적에서 제외" 설정 가능\n• 설정 시 해당 카테고리의 모든 거래가 카드 실적 집계에서 제외됨\n• 계좌이체·대출 상환 전용 카테고리 등에 활용\n• 실적 제외 카테고리는 목록에 "실적제외" 배지로 표시'),
        ('⚙️', '설정', '앱 환경을 설정합니다.\n\n• 🆕 업데이트 내역: 최근 업데이트 내용을 언제든 다시 확인 (이전 버전 기록도 열람 가능)\n• 공지사항: 앱 업데이트 및 안내 확인\n• 🏷️ 카테고리 관리: 지출·수입 카테고리 추가·수정·삭제·순서 변경\n• 포트폴리오: 자산 현황을 PDF로 출력\n• 🔒 보안: 닉네임·비밀번호 변경, 로그아웃, 회원 탈퇴'),
        ('📄', '포트폴리오 PDF', '나의 자산 현황을 PDF 파일로 저장합니다.\n\n• 설정 → 포트폴리오 PDF 출력 버튼 클릭\n• 포함할 항목 선택 후 PDF 출력\n• 미리보기 화면에서 ⬇ PDF 저장 버튼 클릭\n• 모바일: 공유 → 파일로 저장 / PC: 인쇄 → PDF로 저장\n\n자산 구성\n• 현금·예금·적금/청약·투자 항목 포함, 금액 큰 순서대로 정렬\n• 도넛 차트 중앙은 대출 차감 후 순자산 표시\n• 대출 계좌가 있으면 구분선 아래 빨간색으로 별도 표시\n• 이달 수입·지출은 통계에서 제외한 내역을 자동으로 제외\n\n거래내역\n• 기본 비활성화 — 체크 시 최근 30건만 출력'),
    ]
    _help_v_cfg = AppConfig.query.get('help_version')
    if _help_v_cfg is None or _help_v_cfg.value != _help_version:
        HelpItem.query.delete()
        for i, (icon, title, desc) in enumerate(_help_defaults):
            db.session.add(HelpItem(icon=icon, title=title, desc=desc, position=i))
        if _help_v_cfg:
            _help_v_cfg.value = _help_version
        else:
            db.session.add(AppConfig(key='help_version', value=_help_version))
        db.session.commit()

    # one-time fix: reset price_updated_at for 해외주식 so auto-fetch re-runs
    # (previous version stored current_price in KRW; new version stores in USD)
    if AppConfig.query.get('fix_overseas_price_unit') is None:
        from models import Investment as _Inv
        _fixed = _Inv.query.filter_by(itype='해외주식').all()
        for _inv in _fixed:
            _inv.price_updated_at = None
        db.session.add(AppConfig(key='fix_overseas_price_unit', value='done'))
        db.session.commit()

    # seed / upgrade update notice config
    _notice_v = 'ver 2.29'
    _notice = {
        'version': _notice_v,
        'date': '2026년 7월 11일',
        'updates': [
            {'section': '💰 월급', 'items': [
                {'tag': 'new', 'title': '카테고리 순서 드래그 변경', 'desc': '월급 탭 예산 배분 카테고리를 ⠿ 핸들로 드래그하여 순서 변경 — 드래그 중 카드가 실시간으로 따라오는 미리보기 제공'},
            ]},
            {'section': '📅 캘린더', 'items': [
                {'tag': 'new', 'title': '내역 정렬 버튼', 'desc': '월별 내역 목록에 최신순 / 오래된순 정렬 토글 버튼 추가 — 달 이동 시 최신순으로 초기화'},
                {'tag': 'fix', 'title': '수입·지출·총계 통계제외 반영', 'desc': '캘린더 하단 수입/지출/총계 합계에 "통계에서 제외"한 내역이 포함되던 문제 수정'},
            ]},
            {'section': '🎨 UI 개선', 'items': [
                {'tag': 'imp', 'title': '내역 카테고리 이모지 표시', 'desc': '홈·캘린더 탭 내역 목록에서 카테고리 이름 앞에 설정된 이모지 표시'},
                {'tag': 'imp', 'title': '홈·캘린더 내역 스타일 통일', 'desc': '캘린더 탭 날짜 클릭 팝업·월별 목록이 홈 탭과 동일한 스타일(카드배지·이모지·배지)로 표시'},
            ]},
            {'section': '📄 포트폴리오', 'items': [
                {'tag': 'imp', 'title': '자산 구성 순자산 표시', 'desc': '도넛 차트 중앙이 "총 자산" 대신 대출 차감 후 "순자산"으로 표시'},
                {'tag': 'imp', 'title': '자산 구성 통장잔고 제외', 'desc': '포트폴리오 자산 구성 도넛에서 통장잔고 항목 제거 — 예금·적금/청약·투자 중심으로 표시'},
                {'tag': 'imp', 'title': '청약 경과 기간 표시', 'desc': '청약 카드의 D-NaN·null개월 오류 수정 — 시작일 기준 경과 일수(D+X)와 경과 개월수로 정확히 표시'},
            ]},
            {'section': '🔧 버그 수정 · 개선', 'items': [
                {'tag': 'fix', 'title': '통계 청약 추가입금 반영', 'desc': '청약 추가입금이 통계 탭 적금/청약 합계에 반영되지 않던 문제 수정'},
                {'tag': 'fix', 'title': '홈·포트폴리오 통계제외 필터링', 'desc': '홈 탭 카테고리별 지출과 포트폴리오 이달 지출에 통계에서 제외한 내역이 포함되던 문제 수정'},
                {'tag': 'imp', 'title': '자산 구성 대출 표시', 'desc': '통계 탭 자산 구성 "대출/빚" 레이블을 "대출"로 변경'},
            ]},
        ]
    }
    existing = AppConfig.query.get('update_notice')
    if existing is None:
        db.session.add(AppConfig(key='update_notice', value=json.dumps(_notice, ensure_ascii=False)))
    else:
        try:
            cur = json.loads(existing.value)
            if cur.get('version') != _notice_v:
                existing.value = json.dumps(_notice, ensure_ascii=False)
        except Exception:
            existing.value = json.dumps(_notice, ensure_ascii=False)
    db.session.commit()

# ── Auth routes ───────────────────────────────────────────────────────────────

@app.route('/api/me')
def api_me():
    uid = session.get('user_id')
    if not uid:
        return jsonify({'user': None})
    user = User.query.get(uid)
    if not user:
        session.pop('user_id', None)
        return jsonify({'user': None})
    return jsonify({'user': {'id': user.id, 'email': user.email, 'nickname': user.nickname}})

@app.route('/api/register', methods=['POST'])
def api_register():
    data = request.json or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    nickname = data.get('nickname', '').strip()
    if not email or not password or len(password) < 6:
        return jsonify({'error': '이메일과 비밀번호(6자 이상)를 입력하세요'}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({'error': '이미 사용 중인 이메일입니다'}), 400
    user = User(email=email, nickname=nickname or None)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    _seed_user_categories(user.id)
    session.permanent = True
    session['user_id'] = user.id
    return jsonify({'ok': True, 'email': user.email, 'nickname': user.nickname})

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    user = User.query.filter_by(email=email).first()
    if not user or not user.check_password(password):
        return jsonify({'error': '이메일 또는 비밀번호가 올바르지 않습니다'}), 401
    session.permanent = True
    session['user_id'] = user.id
    return jsonify({'ok': True, 'email': user.email, 'nickname': user.nickname})

def send_reset_email(to_email, code):
    mail_user = os.environ.get('MAIL_USER', '')
    mail_pass = os.environ.get('MAIL_PASSWORD', '')
    if not mail_user or not mail_pass:
        raise RuntimeError('이메일 설정이 되어 있지 않습니다')
    msg = MIMEText(f'인증번호: {code}\n\n30분 이내에 입력해주세요.', 'plain', 'utf-8')
    msg['Subject'] = '[나의 가계부] 비밀번호 재설정 인증번호'
    msg['From'] = mail_user
    msg['To'] = to_email
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as s:
        s.login(mail_user, mail_pass)
        s.sendmail(mail_user, to_email, msg.as_string())

@app.route('/api/reset-request', methods=['POST'])
def api_reset_request():
    email = (request.json or {}).get('email', '').strip().lower()
    user = User.query.filter_by(email=email).first()
    if not user:
        return jsonify({'error': '등록된 이메일이 없습니다'}), 404
    code = '%06d' % random.randint(0, 999999)
    user.reset_code = code
    user.reset_expires = datetime.now() + timedelta(minutes=30)
    db.session.commit()
    try:
        send_reset_email(email, code)
    except Exception as e:
        return jsonify({'error': f'이메일 발송에 실패했습니다: {str(e)}'}), 500
    return jsonify({'ok': True})

@app.route('/api/reset-confirm', methods=['POST'])
def api_reset_confirm():
    data = request.json or {}
    email = data.get('email', '').strip().lower()
    code = data.get('code', '')
    new_pw = data.get('password', '')
    user = User.query.filter_by(email=email).first()
    if not user or user.reset_code != code or not user.reset_expires or datetime.now() > user.reset_expires:
        return jsonify({'error': '코드가 잘못되었거나 만료되었습니다 (30분)'}), 400
    if len(new_pw) < 6:
        return jsonify({'error': '비밀번호는 6자 이상이어야 합니다'}), 400
    user.set_password(new_pw)
    user.reset_code = None
    user.reset_expires = None
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.pop('user_id', None)
    return jsonify({'ok': True})

@app.route('/api/update-nickname', methods=['POST'])
@login_required
def api_update_nickname():
    nickname = (request.json or {}).get('nickname', '').strip()
    user = User.query.get(session['user_id'])
    user.nickname = nickname or None
    db.session.commit()
    return jsonify({'ok': True, 'nickname': user.nickname})

@app.route('/api/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.json or {}
    current_pw = data.get('current_password', '')
    new_pw = data.get('new_password', '')
    if not current_pw or not new_pw:
        return jsonify({'error': '비밀번호를 입력하세요'}), 400
    if len(new_pw) < 6:
        return jsonify({'error': '새 비밀번호는 6자 이상이어야 합니다'}), 400
    user = User.query.get(session['user_id'])
    if not user.check_password(current_pw):
        return jsonify({'error': '현재 비밀번호가 올바르지 않습니다'}), 400
    user.set_password(new_pw)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/delete-account', methods=['POST'])
@login_required
def api_delete_account():
    uid = session['user_id']
    from sqlalchemy import text as _text
    with db.engine.connect() as conn:
        conn.execute(_text('DELETE FROM "transaction" WHERE user_id = :uid'), {'uid': uid})
        conn.commit()
    Budget.query.filter_by(user_id=uid).delete()
    Category.query.filter_by(user_id=uid).delete()
    Card.query.filter_by(user_id=uid).delete()
    user = User.query.get(uid)
    db.session.delete(user)
    db.session.commit()
    session.pop('user_id', None)
    return jsonify({'ok': True})

# ── Transaction filter helpers ───────────────────────────────────────────────

def _is_stats_tx(tx, excl_cats=frozenset()):
    return not tx.exclude_stats and tx.category not in excl_cats

def _is_perf_tx(tx, excl_cats=frozenset()):
    return not tx.exclude_perf and tx.category not in excl_cats

# ── Savings helper ───────────────────────────────────────────────────────────

def _savings_stats(s, extra_deposit=0):
    from datetime import date as _date
    today = _date.today()
    if s.stype == '청약':
        try:
            start = datetime.strptime(s.start_date, '%Y-%m-%d').date()
        except Exception:
            start = today
        months_elapsed_auto = max(0, (today.year - start.year) * 12 + (today.month - start.month))
        manual_count = getattr(s, 'manual_count', None)
        months_elapsed = manual_count if manual_count is not None else months_elapsed_auto
        current_paid = s.amount * months_elapsed + extra_deposit
        rate = s.interest_rate or 0
        itype = getattr(s, 'interest_type', '단리') or '단리'
        tax_type = getattr(s, 'tax_type', '비과세') or '비과세'
        n = max(1, months_elapsed)
        r_m = rate / 100 / 12
        if rate > 0:
            if itype == '복리' and r_m > 0:
                interest = int(s.amount * (1 + r_m) * ((1 + r_m) ** n - 1) / r_m) - current_paid
            else:
                interest = int(s.amount * r_m * n * (n + 1) / 2)
        else:
            interest = 0
        if tax_type == '비과세':
            tax = 0
        elif tax_type == '세금우대':
            income_tax = (int(interest * 0.09) // 10) * 10
            tax = income_tax + (int(income_tax * 0.1) // 10) * 10
        elif tax_type.startswith('ISA') or tax_type == 'ISA':
            threshold = 4_000_000 if tax_type == 'ISA(서민형)' else 2_000_000
            taxable = max(0, interest - threshold)
            income_tax = (int(taxable * 0.09) // 10) * 10
            tax = income_tax + (int(income_tax * 0.1) // 10) * 10
        else:
            income_tax = (int(interest * 0.14) // 10) * 10
            tax = income_tax + (int(income_tax * 0.1) // 10) * 10
        interest_after_tax = interest - tax
        return {
            'id': s.id, 'stype': s.stype, 'bank': s.bank, 'name': s.name,
            'amount': s.amount, 'interest_rate': rate, 'interest_type': itype,
            'tax_type': tax_type,
            'start_date': s.start_date, 'end_date': '',
            'months_total': None, 'months_elapsed': months_elapsed,
            'months_elapsed_auto': months_elapsed_auto,
            'progress': 0, 'd_day': None,
            'total_paid': current_paid, 'current_paid': current_paid,
            'interest': interest, 'maturity_amount': current_paid + interest_after_tax,
            'interest_after_tax': interest_after_tax, 'maturity_after_tax': current_paid + interest_after_tax,
            'extra_deposit': extra_deposit,
            'notify_day': getattr(s, 'notify_day', None),
            'auto_tx': bool(getattr(s, 'auto_tx', False)),
            'auto_tx_day': getattr(s, 'auto_tx_day', None),
            'auto_tx_card': getattr(s, 'auto_tx_card', '') or '',
            'manual_count': manual_count,
            'is_paused': bool(getattr(s, 'is_paused', False)),
        }
    try:
        start = datetime.strptime(s.start_date, '%Y-%m-%d').date()
        end = datetime.strptime(s.end_date, '%Y-%m-%d').date()
    except Exception:
        start = today; end = today
    months_total = max(1, (end.year - start.year) * 12 + (end.month - start.month))
    months_elapsed_auto = max(0, min(months_total, (today.year - start.year) * 12 + (today.month - start.month)))
    _manual_count = getattr(s, 'manual_count', None)
    months_elapsed = _manual_count if _manual_count is not None else months_elapsed_auto
    days_total = max(1, (end - start).days)
    days_elapsed = max(0, min(days_total, (today - start).days)) if _manual_count is None else int(_manual_count / months_total * days_total)
    progress = min(100.0, round(days_elapsed / days_total * 100, 1))
    d_day = (end - today).days
    rate = s.interest_rate or 0
    itype = getattr(s, 'interest_type', '단리') or '단리'
    tax_type = getattr(s, 'tax_type', '일반과세') or '일반과세'
    if s.stype == '예금':
        total_paid = s.amount
        current_paid = s.amount
        if itype == '복리':
            interest = int(s.amount * (1 + rate / 100 / 12) ** months_total) - s.amount
        else:
            interest = int(s.amount * rate / 100 * days_total / 365)
        maturity_amount = s.amount + interest
    else:  # 적금, 청약
        total_paid = s.amount * months_total
        current_paid = s.amount * months_elapsed
        bonus_amount = getattr(s, 'bonus_amount', None) or 0
        bonus_total = bonus_amount * months_total if bonus_amount else 0
        bonus_current = bonus_amount * months_elapsed if bonus_amount else 0
        n = months_total
        r_m = rate / 100 / 12
        if itype == '복리' and r_m > 0:
            interest = int(s.amount * (1 + r_m) * ((1 + r_m) ** n - 1) / r_m) - total_paid
        else:
            # 적금 단리: 월납입액 × 월이율 × n(n+1)/2
            interest = int(s.amount * r_m * n * (n + 1) / 2)
        maturity_amount = total_paid + bonus_total + interest
    # 세금: 이자소득세(14%) 원 미만 절사 → 지방소득세 = 이자소득세의 10% 원 미만 절사
    if tax_type == '비과세':
        tax = 0
    elif tax_type == '세금우대':
        income_tax = (int(interest * 0.09) // 10) * 10
        tax = income_tax + (int(income_tax * 0.1) // 10) * 10
    elif tax_type.startswith('ISA') or tax_type == 'ISA':
        threshold = 4_000_000 if tax_type == 'ISA(서민형)' else 2_000_000
        taxable = max(0, interest - threshold)
        income_tax = (int(taxable * 0.09) // 10) * 10
        tax = income_tax + (int(income_tax * 0.1) // 10) * 10
    else:
        income_tax = (int(interest * 0.14) // 10) * 10
        tax = income_tax + (int(income_tax * 0.1) // 10) * 10
    interest_after_tax = interest - tax
    principal = s.amount if s.stype == '예금' else total_paid
    maturity_after_tax = principal + interest_after_tax + (bonus_total if s.stype != '예금' else 0)
    return {
        'id': s.id, 'stype': s.stype, 'bank': s.bank, 'name': s.name,
        'amount': s.amount, 'interest_rate': rate, 'interest_type': itype,
        'tax_type': tax_type,
        'start_date': s.start_date, 'end_date': s.end_date,
        'months_total': months_total, 'months_elapsed': months_elapsed,
        'months_elapsed_auto': months_elapsed_auto,
        'progress': progress, 'd_day': d_day,
        'total_paid': total_paid, 'current_paid': current_paid,
        'interest': interest, 'maturity_amount': maturity_amount,
        'interest_after_tax': interest_after_tax, 'maturity_after_tax': maturity_after_tax,
        'notify_day': getattr(s, 'notify_day', None),
        'auto_tx': bool(getattr(s, 'auto_tx', False)),
        'auto_tx_day': getattr(s, 'auto_tx_day', None),
        'auto_tx_card': getattr(s, 'auto_tx_card', '') or '',
        'manual_count': _manual_count,
        'is_paused': bool(getattr(s, 'is_paused', False)),
        'bonus_amount': getattr(s, 'bonus_amount', None) or 0,
        'bonus_total': bonus_total if s.stype != '예금' else 0,
        'bonus_current': bonus_current if s.stype != '예금' else 0,
    }

_KST = timezone(timedelta(hours=9))

def _last_market_close(market):
    now = datetime.now(_KST)
    if market == 'KR':
        for delta in range(8):
            d = now.date() - timedelta(days=delta)
            if d.weekday() < 5:
                dt = datetime(d.year, d.month, d.day, 15, 30, tzinfo=_KST)
                if dt <= now:
                    return dt
    elif market == 'US':
        for delta in range(8):
            avail_date = now.date() - timedelta(days=delta)
            avail_dt = datetime(avail_date.year, avail_date.month, avail_date.day, 6, 0, tzinfo=_KST)
            if avail_dt > now:
                continue
            us_trade_day = avail_date - timedelta(days=1)
            if us_trade_day.weekday() < 5:
                return avail_dt
    return now - timedelta(days=30)

def _inv_market(inv):
    if inv.itype == '국내주식':
        return 'KR'
    if inv.itype == '해외주식':
        return 'US'
    if inv.itype == 'ETF':
        t = (inv.ticker or '').strip().upper()
        if t.endswith('.KS') or t.endswith('.KQ') or t.replace('.', '').isdigit():
            return 'KR'
        return 'US'
    return None

def _needs_price_update(inv):
    market = _inv_market(inv)
    if market is None or not (inv.ticker or '').strip():
        return False
    if inv.price_updated_at is None:
        return True
    last = inv.price_updated_at.replace(tzinfo=timezone.utc).astimezone(_KST)
    return last < _last_market_close(market)

def _auto_fetch_investment_prices(inv_list):
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from datetime import date, timedelta
    import FinanceDataReader as fdr

    todo = [inv for inv in inv_list if _needs_price_update(inv) and (inv.ticker or '').strip()]
    if not todo:
        return

    start = (date.today() - timedelta(days=7)).strftime('%Y-%m-%d')

    usd_krw = 1380
    try:
        fx = fdr.DataReader('USD/KRW', start)
        if not fx.empty:
            usd_krw = float(fx['Close'].iloc[-1])
    except Exception:
        pass

    def fetch_price(inv):
        try:
            market = _inv_market(inv)
            t = (inv.ticker or '').strip()
            if market == 'KR':
                t_clean = t.replace('.KS', '').replace('.KQ', '')
                df = fdr.DataReader(t_clean, start)
            elif market == 'US':
                df = fdr.DataReader(t, start)
            else:
                return inv.id, None, None
            if df.empty:
                return inv.id, None, None
            price = float(df['Close'].iloc[-1])
            # US stocks: store in USD (same unit as avg_price). KR stocks: store in KRW.
            new_fx = usd_krw if market == 'US' else None
            return inv.id, price, new_fx
        except Exception:
            return inv.id, None, None

    id_map = {inv.id: inv for inv in todo}
    with ThreadPoolExecutor(max_workers=min(3, len(todo))) as ex:
        futures = {ex.submit(fetch_price, inv): inv.id for inv in todo}
        results = {}
        for f in as_completed(futures):
            inv_id, price, new_fx = f.result()
            results[inv_id] = (price, new_fx)

    changed = False
    for inv_id, (price, new_fx) in results.items():
        if price is not None:
            id_map[inv_id].current_price = price
            if new_fx is not None:
                id_map[inv_id].exchange_rate = new_fx
            id_map[inv_id].price_updated_at = datetime.utcnow()
            changed = True
    if changed:
        db.session.commit()

def _investment_stats(inv):
    qty = inv.quantity or 0
    avg = inv.avg_price or 0
    cur = inv.current_price if inv.current_price is not None else avg
    fx = inv.exchange_rate or None
    if inv.itype == '해외주식' and fx:
        purchase_value = int(qty * avg * fx)
        current_value = int(qty * cur * fx)
    else:
        purchase_value = int(qty * avg)
        current_value = int(qty * cur)
    profit = current_value - purchase_value
    profit_pct = round(profit / purchase_value * 100, 2) if purchase_value else 0
    updated_at = None
    if inv.price_updated_at:
        updated_at = inv.price_updated_at.replace(tzinfo=timezone.utc).astimezone(_KST).strftime('%m.%d %H:%M')
    return {
        'id': inv.id, 'itype': inv.itype, 'name': inv.name,
        'ticker': inv.ticker or '', 'quantity': qty,
        'avg_price': avg, 'current_price': cur,
        'exchange_rate': fx,
        'purchase_value': purchase_value, 'current_value': current_value,
        'profit': profit, 'profit_pct': profit_pct,
        'memo': inv.memo or '',
        'account_type': getattr(inv, 'account_type', '일반') or '일반',
        'price_updated_at': updated_at,
    }

# ── JSON API routes ───────────────────────────────────────────────────────────

@app.route('/api/home')
@login_required
def api_home():
    uid = session['user_id']
    current_month = datetime.now().strftime('%Y-%m')
    transactions = Transaction.query.filter_by(user_id=uid).order_by(Transaction.date.desc()).all()
    month_txs = [tx for tx in transactions if tx.date.startswith(current_month)]

    income_total = sum(tx.amount for tx in month_txs if tx.type == 'income' and _is_stats_tx(tx))
    expense_total = sum(tx.amount for tx in month_txs if tx.type == 'expense' and _is_stats_tx(tx))

    budget = Budget.query.filter_by(month=current_month, user_id=uid).first()
    budget_amount = budget.amount if budget else 0

    cards = Card.query.filter_by(user_id=uid).all()
    expense_cats = Category.query.filter_by(user_id=uid, cat_type='expense').order_by(Category.position, Category.id).all()
    income_cats = Category.query.filter_by(user_id=uid, cat_type='income').order_by(Category.position, Category.id).all()
    excl_cats = {c.name for c in expense_cats if c.exclude_perf}
    excl_stat_cats = {c.name for c in (expense_cats + income_cats) if c.exclude_stats}
    card_stats = []
    for card in cards:
        spent = sum(tx.amount for tx in month_txs
                    if tx.type == 'expense' and tx.card == card.name
                    and _is_perf_tx(tx, excl_cats))
        card_stats.append({
            'name': card.name,
            'target': card.monthly_target,
            'spent': spent,
            'percent': min(int(spent / card.monthly_target * 100), 100) if card.monthly_target > 0 else 0,
            'tier1': card.tier1 or 20, 'tier2': card.tier2 or 50, 'tier3': card.tier3 or 80,
            'is_loan': (card.account_balance or 0) < 0,
        })
    emoji_map = {c.name: c.icon for c in expense_cats + income_cats}

    category_totals = defaultdict(int)
    for tx in month_txs:
        if tx.type == 'expense' and _is_stats_tx(tx, excl_stat_cats):
            category_totals[tx.category] += tx.amount
    category_totals = dict(sorted(category_totals.items(), key=lambda x: x[1], reverse=True))

    return jsonify({
        'transactions': [{'id': tx.id, 'date': tx.date, 'type': tx.type, 'category': tx.category,
                          'description': tx.description or '', 'amount': tx.amount, 'card': tx.card or '',
                          'exclude_perf': bool(tx.exclude_perf), 'exclude_stats': bool(tx.exclude_stats)} for tx in month_txs],
        'income_total': income_total,
        'expense_total': expense_total,
        'balance': income_total - expense_total,
        'budget_amount': budget_amount,
        'remaining': budget_amount - expense_total,
        'card_stats': card_stats,
        'card_list': [{'id': c.id, 'name': c.name, 'is_loan': (c.account_balance or 0) < 0} for c in cards],
        'expense_cats': [[c.name, c.icon] for c in expense_cats],
        'income_cats': [[c.name, c.icon] for c in income_cats],
        'emoji_map': emoji_map,
        'category_totals': category_totals,
        'excl_cat_names': list(excl_cats),
        'excl_stat_cat_names': list(excl_stat_cats),
    })

@app.route('/api/transactions', methods=['POST'])
@login_required
def api_add_transaction():
    uid = session['user_id']
    data = request.json or {}
    tx = Transaction(
        date=data['date'], type=data['type'], category=data['category'],
        description=data.get('description', ''), amount=int(data['amount']),
        card=data.get('card') or None,
        exclude_perf=bool(data.get('exclude_perf', False)),
        exclude_stats=bool(data.get('exclude_stats', False)),
        user_id=uid,
    )
    db.session.add(tx)
    db.session.commit()
    return jsonify({'ok': True, 'id': tx.id})

@app.route('/api/transactions/<int:tx_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_transaction(tx_id):
    uid = session['user_id']
    tx = Transaction.query.filter_by(id=tx_id, user_id=uid).first_or_404()
    if request.method == 'DELETE':
        db.session.delete(tx)
        db.session.commit()
        return jsonify({'ok': True})
    if request.method == 'PUT':
        data = request.json or {}
        tx.date = data.get('date', tx.date)
        tx.type = data.get('type', tx.type)
        tx.category = data.get('category', tx.category)
        tx.description = data.get('description', tx.description)
        tx.amount = int(data.get('amount', tx.amount))
        tx.card = data.get('card') or None
        if 'exclude_perf' in data:
            tx.exclude_perf = bool(data['exclude_perf'])
        if 'exclude_stats' in data:
            tx.exclude_stats = bool(data['exclude_stats'])
        db.session.commit()
        return jsonify({'ok': True})
    expense_cats = Category.query.filter_by(user_id=uid, cat_type='expense').order_by(Category.position, Category.id).all()
    income_cats = Category.query.filter_by(user_id=uid, cat_type='income').order_by(Category.position, Category.id).all()
    return jsonify({
        'transaction': {'id': tx.id, 'date': tx.date, 'type': tx.type, 'category': tx.category,
                        'description': tx.description or '', 'amount': tx.amount, 'card': tx.card or '',
                        'exclude_perf': bool(tx.exclude_perf), 'exclude_stats': bool(tx.exclude_stats)},
        'expense_cats': [[c.name, c.icon] for c in expense_cats],
        'income_cats': [[c.name, c.icon] for c in income_cats],
        'card_list': [{'id': c.id, 'name': c.name, 'is_loan': (c.account_balance or 0) < 0} for c in Card.query.filter_by(user_id=uid).all()],
    })

@app.route('/api/cards', methods=['GET', 'POST'])
@login_required
def api_cards():
    uid = session['user_id']
    if request.method == 'POST':
        data = request.json or {}
        db.session.add(Card(
            name=data['name'], monthly_target=int(data.get('target', 0)),
            tier1=int(data.get('tier1', 20)), tier2=int(data.get('tier2', 50)), tier3=int(data.get('tier3', 80)),
            account_balance=int(data.get('account_balance', 0)),
            url=data.get('url') or None,
            user_id=uid,
            linked_account_id=data.get('linked_account_id') or None,
        ))
        db.session.commit()
        return jsonify({'ok': True})
    current_month = datetime.now().strftime('%Y-%m')
    month_txs = Transaction.query.filter_by(user_id=uid).filter(Transaction.date.like(f'{current_month}%')).all()
    cards = Card.query.filter_by(user_id=uid).all()
    excl_cats_cards = {c.name for c in Category.query.filter_by(user_id=uid, exclude_perf=True).all()}
    stats = {}
    for card in cards:
        spent = sum(tx.amount for tx in month_txs
                    if tx.type == 'expense' and tx.card == card.name
                    and _is_perf_tx(tx, excl_cats_cards))
        stats[card.id] = {
            'spent': spent,
            'percent': min(int(spent / card.monthly_target * 100), 100) if card.monthly_target > 0 else 0,
        }
    return jsonify({
        'cards': [{'id': c.id, 'name': c.name, 'target': c.monthly_target, 'url': c.url or '',
                   'tier1': c.tier1 or 20, 'tier2': c.tier2 or 50, 'tier3': c.tier3 or 80,
                   'account_balance': c.account_balance or 0, 'linked_account_id': c.linked_account_id} for c in cards],
        'stats': stats,
    })

@app.route('/api/cards/<int:card_id>', methods=['PUT', 'DELETE'])
@login_required
def api_card(card_id):
    uid = session['user_id']
    card = Card.query.filter_by(id=card_id, user_id=uid).first_or_404()
    if request.method == 'DELETE':
        db.session.delete(card)
        db.session.commit()
        return jsonify({'ok': True})
    data = request.json or {}
    card.name = data.get('name', card.name)
    card.monthly_target = int(data.get('target', card.monthly_target))
    card.tier1 = int(data.get('tier1', card.tier1 or 20))
    card.tier2 = int(data.get('tier2', card.tier2 or 50))
    card.tier3 = int(data.get('tier3', card.tier3 or 80))
    if 'account_balance' in data:
        card.account_balance = int(data['account_balance'])
    if 'url' in data:
        card.url = data['url'] or None
    if 'linked_account_id' in data:
        card.linked_account_id = data['linked_account_id'] or None
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/calendar')
@login_required
def api_calendar():
    uid = session['user_id']
    now = datetime.now()
    month = request.args.get('month', now.strftime('%Y-%m'))
    transactions = Transaction.query.filter_by(user_id=uid).filter(Transaction.date.like(f'{month}%')).all()

    cats = Category.query.filter_by(user_id=uid).order_by(Category.position, Category.id).all()
    emoji_map = {c.name: c.icon for c in cats}
    excl_stat_cats = {c.name for c in cats if c.exclude_stats}

    day_totals = defaultdict(lambda: {'expense': 0, 'income': 0})
    day_transactions = defaultdict(list)
    for tx in transactions:
        day_totals[tx.date][tx.type] += tx.amount
        day_transactions[tx.date].append({
            'id': tx.id, 'type': tx.type, 'category': tx.category,
            'description': tx.description or '', 'amount': tx.amount, 'card': tx.card or '',
            'exclude_perf': bool(tx.exclude_perf), 'exclude_stats': bool(tx.exclude_stats),
        })

    return jsonify({
        'day_totals': {k: dict(v) for k, v in day_totals.items()},
        'day_transactions': dict(day_transactions),
        'income_total': sum(tx.amount for tx in transactions if tx.type == 'income' and _is_stats_tx(tx, excl_stat_cats)),
        'expense_total': sum(tx.amount for tx in transactions if tx.type == 'expense' and _is_stats_tx(tx, excl_stat_cats)),
        'emoji_map': emoji_map,
    })

@app.route('/api/stats')
@login_required
def api_stats():
    uid = session['user_id']
    now = datetime.now()
    month = request.args.get('month', now.strftime('%Y-%m'))

    cats = Category.query.filter_by(user_id=uid).order_by(Category.position, Category.id).all()
    emoji_map = {c.name: c.icon for c in cats}
    icon_map = {c.name: c.icon for c in cats}

    expense_txs_all = Transaction.query.filter_by(user_id=uid).filter(
        Transaction.type == 'expense', Transaction.date.like(f'{month}%')).all()
    income_txs_all = Transaction.query.filter_by(user_id=uid).filter(
        Transaction.type == 'income', Transaction.date.like(f'{month}%')).all()
    excl_stat_cats_stats = {c.name for c in cats if c.exclude_stats}
    expense_txs = [tx for tx in expense_txs_all if _is_stats_tx(tx, excl_stat_cats_stats)]
    income_txs = [tx for tx in income_txs_all if _is_stats_tx(tx, excl_stat_cats_stats)]

    def cat_totals(txs):
        totals = defaultdict(int)
        for tx in txs:
            totals[tx.category] += tx.amount
        return sorted([{'name': k, 'amount': v, 'icon': icon_map.get(k, '📦')} for k, v in totals.items()], key=lambda x: x['amount'], reverse=True)

    bar_from_raw = request.args.get('bar_from')
    bar_to_raw = request.args.get('bar_to')
    if bar_from_raw and bar_to_raw:
        bf_y, bf_m = int(bar_from_raw[:4]), int(bar_from_raw[5:])
        bt_y, bt_m = int(bar_to_raw[:4]), int(bar_to_raw[5:])
    else:
        bt_y, bt_m = now.year, now.month
        bf_m = bt_m - 5; bf_y = bt_y
        while bf_m <= 0: bf_m += 12; bf_y -= 1

    bar_months = []
    cy2, cm2 = bf_y, bf_m
    while (cy2, cm2) <= (bt_y, bt_m):
        bar_months.append(f'{cy2}-{cm2:02d}')
        cm2 += 1
        if cm2 > 12: cm2 = 1; cy2 += 1

    monthly = []
    for mo in bar_months:
        e = Transaction.query.filter_by(user_id=uid).filter(
            Transaction.type == 'expense', Transaction.date.like(f'{mo}%')).all()
        inc = Transaction.query.filter_by(user_id=uid).filter(
            Transaction.type == 'income', Transaction.date.like(f'{mo}%')).all()
        monthly.append({
            'month': mo,
            'expense': sum(t.amount for t in e if _is_stats_tx(t, excl_stat_cats_stats)),
            'income': sum(t.amount for t in inc if _is_stats_tx(t, excl_stat_cats_stats)),
        })

    cards = Card.query.filter_by(user_id=uid).all()
    excl_cats_stats = {c.name for c in cats if c.exclude_perf}
    card_monthly_trend = {}
    for card in cards:
        trend = []
        for mo in bar_months:
            mo_txs = Transaction.query.filter_by(user_id=uid).filter(
                Transaction.type == 'expense',
                Transaction.date.like(f'{mo}%'),
                Transaction.card == card.name,
            ).all()
            amt = sum(tx.amount for tx in mo_txs if _is_perf_tx(tx, excl_cats_stats))
            trend.append(amt)
        card_monthly_trend[card.name] = trend

    card_monthly = []
    for card in cards:
        spent = sum(tx.amount for tx in expense_txs
                    if tx.card == card.name and _is_perf_tx(tx, excl_cats_stats))
        if spent > 0:
            card_monthly.append({'name': card.name, 'spent': spent})
    card_monthly.sort(key=lambda x: x['spent'], reverse=True)

    # 총 자산 추이 (기간 선택 가능, 최대 6개월)
    import calendar as _cal
    from datetime import date as _date
    all_txs_ever = Transaction.query.filter_by(user_id=uid).all()
    savings_list = Savings.query.filter_by(user_id=uid).all()

    tf_raw = request.args.get('trend_from')
    tt_raw = request.args.get('trend_to')
    try:
        tt_y, tt_m = map(int, tt_raw.split('-'))
    except Exception:
        tt_y, tt_m = now.year, now.month
    try:
        tf_y, tf_m = map(int, tf_raw.split('-'))
    except Exception:
        tf_y, tf_m = tt_y, tt_m - 5
        while tf_m <= 0: tf_m += 12; tf_y -= 1
    # 최대 6개월 강제
    total_months = (tt_y - tf_y) * 12 + (tt_m - tf_m) + 1
    if total_months > 6:
        tf_m = tt_m - 5; tf_y = tt_y
        while tf_m <= 0: tf_m += 12; tf_y -= 1
    if total_months < 1:
        tf_y, tf_m = tt_y, tt_m

    # 투자 현재 평가액
    all_savings_deposits = SavingsDeposit.query.filter_by(user_id=uid).all()
    extra_deposits_stats = {}
    for dep in all_savings_deposits:
        extra_deposits_stats[dep.savings_id] = extra_deposits_stats.get(dep.savings_id, 0) + dep.amount

    inv_list = Investment.query.filter_by(user_id=uid).all()
    inv_by_type = {}
    for inv in inv_list:
        price = (inv.current_price if inv.current_price is not None else inv.avg_price) or 0
        fx = (inv.exchange_rate or 1) if inv.itype == '해외주식' else 1
        val = int((inv.quantity or 0) * price * fx)
        inv_by_type[inv.itype] = inv_by_type.get(inv.itype, 0) + val
    inv_total_now = sum(inv_by_type.values())

    # 포트폴리오 구성 (현재 스냅샷) — 현금 카드 제외
    def _is_cash_card(c):
        return '현금' in c.name or '지갑' in c.name
    card_pos = sum(c.account_balance or 0 for c in cards if not _is_cash_card(c) and (c.account_balance or 0) > 0)
    loan_total = sum(c.account_balance or 0 for c in cards if not _is_cash_card(c) and (c.account_balance or 0) < 0)
    cash_total = sum(c.account_balance or 0 for c in cards if _is_cash_card(c) and (c.account_balance or 0) > 0)
    card_balance_now = card_pos + loan_total  # 순 통장잔고 (대출 차감)
    card_balance_all = sum(c.account_balance or 0 for c in cards)  # 자산 추이용 (현금 포함)
    deposit_total = sum(s.amount for s in savings_list if s.stype == '예금')
    installment_total = sum(_savings_stats(s, extra_deposits_stats.get(s.id, 0))['current_paid'] for s in savings_list if s.stype in ('적금', '청약'))
    total_assets_now = card_balance_now + cash_total + deposit_total + installment_total + inv_total_now  # 순자산 (현금·대출 반영)

    portfolio_breakdown = []
    if card_pos > 0:
        portfolio_breakdown.append({'label': '통장잔고', 'value': card_pos})
    if cash_total > 0:
        portfolio_breakdown.append({'label': '현금', 'value': cash_total})
    if deposit_total > 0:
        portfolio_breakdown.append({'label': '예금', 'value': deposit_total})
    if installment_total > 0:
        portfolio_breakdown.append({'label': '적금/청약', 'value': installment_total})
    for k, v in inv_by_type.items():
        if v > 0:
            portfolio_breakdown.append({'label': k, 'value': v})
    if loan_total < 0:
        portfolio_breakdown.append({'label': '대출', 'value': loan_total})

    card_initial = card_balance_all
    asset_trend = []
    cy, cm = tf_y, tf_m
    while (cy, cm) <= (tt_y, tt_m):
        if (cy, cm) == (tt_y, tt_m):
            bd = {'통장잔고': card_pos, '현금': cash_total, '예금': deposit_total, '적금/청약': installment_total}
            bd.update({k: v for k, v in inv_by_type.items()})
            if loan_total < 0:
                bd['대출'] = loan_total
            trend_total = card_balance_all + deposit_total + installment_total + inv_total_now
            asset_trend.append({'month': f'{cy}-{cm:02d}', 'assets': trend_total, 'breakdown': bd})
        else:
            last_day = _cal.monthrange(cy, cm)[1]
            mo_end = f'{cy}-{cm:02d}-{last_day:02d}'
            inc = sum(tx.amount for tx in all_txs_ever if tx.type == 'income' and tx.date <= mo_end)
            exp = sum(tx.amount for tx in all_txs_ever if tx.type == 'expense' and tx.date <= mo_end)
            card_bal = card_initial + inc - exp
            dep_bal = 0
            inst_bal = 0
            mo_end_date = _date(cy, cm, last_day)
            for s in savings_list:
                if s.start_date > mo_end: continue
                start = datetime.strptime(s.start_date, '%Y-%m-%d').date()
                if s.stype == '청약':
                    me = max(0, (mo_end_date.year - start.year) * 12 + (mo_end_date.month - start.month))
                    extra_dep_mo = sum(d.amount for d in all_savings_deposits if d.savings_id == s.id and d.date <= mo_end)
                    inst_bal += s.amount * me + extra_dep_mo
                elif s.stype == '예금':
                    dep_bal += s.amount
                else:
                    end_d = datetime.strptime(s.end_date, '%Y-%m-%d').date()
                    mt = max(1, (end_d.year - start.year) * 12 + (end_d.month - start.month))
                    me = max(0, min(mt, (mo_end_date.year - start.year) * 12 + (mo_end_date.month - start.month)))
                    inst_bal += s.amount * me
            bd = {'통장잔고': card_bal, '예금': dep_bal, '적금/청약': inst_bal}
            bd.update({k: v for k, v in inv_by_type.items()})
            asset_trend.append({'month': f'{cy}-{cm:02d}', 'assets': card_bal + dep_bal + inst_bal + inv_total_now, 'breakdown': bd})
        cm += 1
        if cm > 12: cm = 1; cy += 1

    first_tx = Transaction.query.filter_by(user_id=uid).order_by(Transaction.date.asc()).first()
    first_month = first_tx.date[:7] if first_tx else month

    return jsonify({
        'expense_cats': cat_totals(expense_txs),
        'income_cats': cat_totals(income_txs),
        'monthly': monthly,
        'emoji_map': emoji_map,
        'card_list': [c.name for c in cards],
        'card_monthly_trend': card_monthly_trend,
        'card_monthly': card_monthly,
        'asset_trend': asset_trend,
        'portfolio_breakdown': portfolio_breakdown,
        'total_assets': total_assets_now,
        'first_month': first_month,
    })

@app.route('/api/portfolio-pdf')
@login_required
def api_portfolio_pdf():
    import math as _math
    import calendar as _cal
    from datetime import date as _date
    uid = session['user_id']
    user_obj = User.query.get(uid)
    today = _date.today()
    current_month = today.strftime('%Y-%m')
    date_str = today.strftime('%Y년 %m월 %d일')
    name_display = user_obj.nickname or user_obj.email

    all_txs = Transaction.query.filter_by(user_id=uid).order_by(Transaction.date.desc()).all()
    month_txs = [tx for tx in all_txs if tx.date.startswith(current_month)]
    income_mo = sum(tx.amount for tx in month_txs if tx.type == 'income')
    expense_mo = sum(tx.amount for tx in month_txs if tx.type == 'expense')

    cards = Card.query.filter_by(user_id=uid).all()
    savings_list = Savings.query.filter_by(user_id=uid).all()
    inv_list = Investment.query.filter_by(user_id=uid).all()

    excl_cats_report = {c.name for c in Category.query.filter_by(user_id=uid, exclude_perf=True).all()}
    excl_stat_cats_report = {c.name for c in Category.query.filter_by(user_id=uid, exclude_stats=True).all()}
    card_stats = []
    for card in cards:
        card_txs = [tx for tx in all_txs if tx.card == card.name]
        c_inc = sum(tx.amount for tx in card_txs if tx.type == 'income')
        c_exp = sum(tx.amount for tx in card_txs if tx.type == 'expense')
        initial = card.account_balance or 0
        balance = initial + c_inc - c_exp
        spent = sum(tx.amount for tx in card_txs
                    if tx.type == 'expense' and tx.date.startswith(current_month)
                    and _is_perf_tx(tx, excl_cats_report) and _is_stats_tx(tx, excl_stat_cats_report))
        percent = min(int(spent / card.monthly_target * 100), 100) if card.monthly_target > 0 else 0
        card_stats.append({'name': card.name, 'initial_balance': initial, 'balance': balance,
                           'spent': spent, 'target': card.monthly_target, 'percent': percent})

    extra_deposits_pdf = {}
    for dep in SavingsDeposit.query.filter_by(user_id=uid).all():
        extra_deposits_pdf[dep.savings_id] = extra_deposits_pdf.get(dep.savings_id, 0) + dep.amount
    sav_stats = [_savings_stats(s, extra_deposits_pdf.get(s.id, 0)) for s in savings_list]
    investments = [_investment_stats(i) for i in inv_list]
    inv_total = sum(i['current_value'] for i in investments)
    inv_gain_total = sum(i['profit'] for i in investments)
    inv_cost_total = sum(i['purchase_value'] for i in investments)
    inv_return_rate = round(inv_gain_total / inv_cost_total * 100, 2) if inv_cost_total else 0
    net_worth = sum(c['balance'] for c in card_stats) + sum(s['current_paid'] for s in sav_stats) + inv_total

    card_bal_total = sum(c['balance'] for c in card_stats)
    deposit_total = sum(s['amount'] for s in sav_stats if s['stype'] == '예금')
    install_total = sum(s['current_paid'] for s in sav_stats if s['stype'] == '적금')
    portfolio = []
    if card_bal_total > 0: portfolio.append(('카드잔고', card_bal_total))
    if deposit_total > 0: portfolio.append(('예금', deposit_total))
    if install_total > 0: portfolio.append(('적금', install_total))
    inv_by_type = {}
    for inv in investments:
        inv_by_type[inv['itype']] = inv_by_type.get(inv['itype'], 0) + inv['current_value']
    for k, v in inv_by_type.items():
        if v > 0: portfolio.append((k, v))
    total_assets = sum(v for _, v in portfolio)

    tt_y, tt_m = today.year, today.month
    tf_m = tt_m - 5; tf_y = tt_y
    while tf_m <= 0: tf_m += 12; tf_y -= 1
    card_initial = sum(c.account_balance or 0 for c in cards)
    asset_trend = []
    cy, cm_i = tf_y, tf_m
    while (cy, cm_i) <= (tt_y, tt_m):
        last_day = _cal.monthrange(cy, cm_i)[1]
        mo_end = f'{cy}-{cm_i:02d}-{last_day:02d}'
        inc = sum(tx.amount for tx in all_txs if tx.type == 'income' and tx.date <= mo_end)
        exp = sum(tx.amount for tx in all_txs if tx.type == 'expense' and tx.date <= mo_end)
        card_bal = card_initial + inc - exp
        sav_bal = 0
        mo_end_date = _date(cy, cm_i, last_day)
        for s in savings_list:
            if s.start_date > mo_end: continue
            start_d = datetime.strptime(s.start_date, '%Y-%m-%d').date()
            if s.stype == '청약':
                me = max(0, (mo_end_date.year - start_d.year) * 12 + (mo_end_date.month - start_d.month))
                sav_bal += s.amount * me
            elif s.stype == '예금':
                sav_bal += s.amount
            else:
                end_d = datetime.strptime(s.end_date, '%Y-%m-%d').date()
                mt = max(1, (end_d.year - start_d.year) * 12 + (end_d.month - start_d.month))
                me = max(0, min(mt, (mo_end_date.year - start_d.year) * 12 + (mo_end_date.month - start_d.month)))
                sav_bal += s.amount * me
        asset_trend.append({'month': f'{cy}-{cm_i:02d}', 'assets': card_bal + sav_bal})
        cm_i += 1
        if cm_i > 12: cm_i = 1; cy += 1

    def fmt(n): return f"{int(n):,}"
    def fmt_short(n):
        n = int(n)
        if abs(n) >= 100000000: return f"{n/100000000:.1f}억원"
        if abs(n) >= 10000: return f"{n/10000:.0f}만원"
        return f"{n:,}원"

    COLORS = ['#b088f9', '#7baff0', '#4BC0C0', '#FF6384', '#FF9F40', '#FFCE56', '#9966FF']

    def make_donut(items, clrs, size=180):
        if not items: return ''
        total = sum(v for _, v in items)
        if not total: return ''
        cx = cy2 = size / 2
        r = size / 2 - 14
        ir = r * 0.58
        angle = -_math.pi / 2
        paths = []
        for i, (label, value) in enumerate(items):
            sweep = min(value / total * 2 * _math.pi, 2 * _math.pi - 0.001)
            x1 = cx + r * _math.cos(angle); y1 = cy2 + r * _math.sin(angle)
            x2 = cx + r * _math.cos(angle + sweep); y2 = cy2 + r * _math.sin(angle + sweep)
            x3 = cx + ir * _math.cos(angle + sweep); y3 = cy2 + ir * _math.sin(angle + sweep)
            x4 = cx + ir * _math.cos(angle); y4 = cy2 + ir * _math.sin(angle)
            la = 1 if sweep > _math.pi else 0
            c = clrs[i % len(clrs)]
            d = f"M{x1:.1f},{y1:.1f} A{r:.1f},{r:.1f} 0 {la},1 {x2:.1f},{y2:.1f} L{x3:.1f},{y3:.1f} A{ir:.1f},{ir:.1f} 0 {la},0 {x4:.1f},{y4:.1f}Z"
            paths.append(f'<path d="{d}" fill="{c}" stroke="#fff" stroke-width="2"/>')
            angle += sweep
        lbl = fmt_short(total)
        center = (f'<text x="{cx:.0f}" y="{cy2-7:.0f}" text-anchor="middle" font-size="13" font-weight="bold" fill="#333">{lbl}</text>'
                  f'<text x="{cx:.0f}" y="{cy2+11:.0f}" text-anchor="middle" font-size="10" fill="#888">총 자산</text>')
        return f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}">{"".join(paths)}{center}</svg>'

    donut_svg = make_donut(portfolio, COLORS)
    legend_html = ''
    for i, (label, value) in enumerate(portfolio):
        pct = value / total_assets * 100 if total_assets else 0
        color = COLORS[i % len(COLORS)]
        legend_html += (
            '<div style="margin-bottom:10px">'
            '<div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">'
            f'<svg width="12" height="12" style="flex-shrink:0;vertical-align:middle"><rect width="12" height="12" rx="3" fill="{color}"/></svg>'
            f'<span style="flex:1;font-size:12px">{label}</span>'
            f'<span style="font-size:12px;font-weight:700">{fmt(value)}원</span>'
            f'<span style="font-size:11px;color:#aaa;width:38px;text-align:right">{pct:.1f}%</span>'
            '</div>'
            f'<svg width="100%" height="8" style="display:block;border-radius:4px;overflow:hidden">'
            f'<rect width="100%" height="8" rx="4" fill="#f0f0f0"/>'
            f'<rect width="{pct:.1f}%" height="8" rx="4" fill="{color}"/>'
            '</svg>'
            '</div>'
        )

    PDF_BANK_LOGOS = [
        ('신한', '/static/cards/sinhanbank.png'), ('KB', '/static/cards/kbbank.png'),
        ('국민', '/static/cards/kbbank.png'), ('농협', '/static/cards/nhbank.png'),
        ('NH', '/static/cards/nhbank.png'), ('하나', '/static/cards/hanabank.png'),
        ('우리', '/static/cards/wooribank.png'), ('기업', '/static/cards/ibkbank.png'),
        ('IBK', '/static/cards/ibkbank.png'), ('카카오', '/static/cards/kakaobank.png'),
        ('토스', '/static/cards/tossbank.png'), ('케이뱅크', '/static/cards/kbank.png'),
        ('K뱅크', '/static/cards/kbank.png'), ('SC', '/static/cards/scbank.png'),
        ('제일', '/static/cards/scbank.png'), ('씨티', '/static/cards/citibank.png'),
        ('iM', '/static/cards/imbank.png'), ('IM', '/static/cards/imbank.png'),
        ('수협', '/static/cards/suhyupbank.png'), ('KDB', '/static/cards/kdbbank.png'),
        ('산업', '/static/cards/kdbbank.png'), ('BNK', '/static/cards/bnkbank.png'),
        ('부산', '/static/cards/bnkbank.png'), ('우체국', '/static/cards/epostbank.png'),
        ('SBI', '/static/cards/sbibank.png'), ('신협', '/static/cards/cubank.png'),
    ]
    def bank_logo_tag(name, size=32):
        if not name: return ''
        for k, url in PDF_BANK_LOGOS:
            if k in name:
                return (f'<img src="{url}" style="width:{size}px;height:{size}px;'
                        f'object-fit:contain;border-radius:8px" '
                        f'onerror="this.style.display=\'none\'" />')
        return ''

    def dday(end_str):
        if not end_str:
            return '—'
        try:
            diff = (_date(*map(int, end_str.split('-'))) - today).days
        except Exception:
            return '—'
        if diff < 0: return f'D+{abs(diff)}'
        if diff == 0: return 'D-Day'
        return f'D-{diff}'

    def card_panel(c):
        tgt = c['target']
        target_html = ''
        if tgt:
            pct_cap = min(c['percent'], 100)
            target_html = (
                '<div class="ig">'
                f'<div class="ic"><div class="l">월 예산</div><div class="v">{fmt(tgt)}원</div></div>'
                f'<div class="ic"><div class="l">이달 실적</div><div class="v">{fmt(c["spent"])}원</div></div>'
                f'<div class="ic"><div class="l">달성률</div><div class="v" style="color:#b088f9">{c["percent"]}%</div></div>'
                '</div>'
                '<div style="margin-top:4px">'
                f'<div class="pb"><div class="pf" style="width:{pct_cap}%;background:linear-gradient(90deg,#b088f9,#7baff0)"></div></div>'
                f'<div class="pl"><span>0원</span><span>{fmt(tgt)}원</span></div>'
                '</div>'
            )
        logo = bank_logo_tag(c['name'])
        return (
            '<div class="cp">'
            f'<div class="cn2" style="display:flex;align-items:center;gap:10px">{logo}<span>{c["name"]}</span></div>'
            '<div class="ig">'
            f'<div class="ic"><div class="l">초기 잔고</div><div class="v">{fmt(c["initial_balance"])}원</div></div>'
            f'<div class="ic"><div class="l">이달 지출</div><div class="v ce">{fmt(c["spent"])}원</div></div>'
            f'<div class="ic"><div class="l">현재 잔고</div><div class="v">{fmt(c["balance"])}원</div></div>'
            '</div>' + target_html + '</div>'
        )

    def sav_panel(s):
        amt_label = '예치금액' if s['stype'] == '예금' else '월 납입액'
        badge_cls = 'by' if s['stype'] == '예금' else 'bj'
        sav_logo = bank_logo_tag(s['bank'] or '', size=28)
        return (
            '<div class="sp"><div class="sh">'
            f'<div class="sn" style="display:flex;align-items:center;gap:8px">{sav_logo}<span>{s["bank"] or ""}</span>&nbsp;<span class="badge {badge_cls}">{s["stype"]}</span></div>'
            f'<div class="dd">{dday(s["end_date"])}</div></div>'
            '<div class="sg2">'
            f'<div class="ic"><div class="l">{amt_label}</div><div class="v">{fmt(s["amount"])}원</div></div>'
            f'<div class="ic"><div class="l">연 이율</div><div class="v">{s["interest_rate"]}%</div></div>'
            f'<div class="ic"><div class="l">기간</div><div class="v">{s["months_total"]}개월</div></div>'
            '</div><div class="sg2">'
            f'<div class="ic"><div class="l">예상 이자</div><div class="v ci">+{fmt(s["interest"])}원</div></div>'
            f'<div class="ic"><div class="l">만기 수령</div><div class="v ci">{fmt(s["maturity_amount"])}원</div></div>'
            f'<div class="ic"><div class="l">만기일</div><div class="v">{s["end_date"]}</div></div>'
            '</div><div style="margin-top:4px">'
            f'<div class="pb"><div class="pf" style="width:{s["progress"]}%;background:linear-gradient(90deg,#b088f9,#7baff0)"></div></div>'
            f'<div class="pl"><span>{s["start_date"]}</span><span>{s["end_date"]}</span></div>'
            '</div></div>'
        )

    # 자산 추이 테이블
    trend_rows = ''
    for i, t in enumerate(asset_trend):
        prev_assets = asset_trend[i-1]['assets'] if i > 0 else None
        chg = t['assets'] - prev_assets if prev_assets is not None else None
        chg_html = ''
        if chg is not None:
            sign = '+' if chg >= 0 else ''
            col = '#198754' if chg >= 0 else '#dc3545'
            arr = '▲' if chg >= 0 else '▼'
            col = '#dc3545' if chg >= 0 else '#0d6efd'
            if prev_assets:
                pct_str = f' ({sign}{chg / prev_assets * 100:.1f}%)'
            elif chg == 0:
                pct_str = ' (+0.0%)'
            else:
                pct_str = ' (신규)'
            chg_html = f'<span style="color:{col};font-size:11px">{arr} {sign}{fmt(chg)}원{pct_str}</span>'
        trend_rows += f'<tr><td style="font-weight:600">{int(t["month"][5:])}월 ({t["month"]})</td><td style="text-align:right;font-weight:700">{fmt(t["assets"])}원</td><td style="text-align:right">{chg_html}</td></tr>'

    # 자산 추이 점선 SVG 라인 차트
    if asset_trend:
        n = len(asset_trend)
        min_val = min(t['assets'] for t in asset_trend)
        max_val = max(t['assets'] for t in asset_trend) or 1
        val_range = max_val - min_val or 1
        W, H, PX, PY = 500, 80, 28, 10
        cw, ch = W - PX * 2, H - PY * 2
        pts = []
        for i, t in enumerate(asset_trend):
            x = PX + (i / (n - 1) * cw if n > 1 else cw / 2)
            y = PY + ch - ((t['assets'] - min_val) / val_range * ch)
            pts.append((x, y, t))
        path_d = ' '.join(f'{"M" if i == 0 else "L"}{x:.1f},{y:.1f}' for i, (x, y, _) in enumerate(pts))
        circles = ''.join(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4" fill="#b088f9" stroke="white" stroke-width="2"/>' for x, y, _ in pts)
        labels = ''.join(f'<text x="{x:.1f}" y="{H + 14}" text-anchor="middle" font-size="9" fill="#888">{int(t["month"][5:])}월</text>' for x, _, t in pts)
        bar_chart = f'<svg width="100%" viewBox="0 0 {W} {H + 20}" style="margin:10px 0"><path d="{path_d}" fill="none" stroke="#b088f9" stroke-width="2" stroke-dasharray="5,4"/>{circles}{labels}</svg>'
    else:
        bar_chart = ''

    cards_html = ''.join(card_panel(c) for c in card_stats) if card_stats else '<div class="empty">등록된 카드/계좌가 없습니다</div>'
    savings_html = ''.join(sav_panel(s) for s in sav_stats) if sav_stats else '<div class="empty">등록된 예적금이 없습니다</div>'

    sav_summary_html = ''
    if sav_stats:
        sav_summary_html = (
            '<div class="sg" style="margin-top:12px">'
            f'<div class="sc"><div class="l">총 예치금</div><div class="v cn">{fmt(sum(s["amount"] for s in sav_stats))}원</div></div>'
            f'<div class="sc"><div class="l">총 예상 이자</div><div class="v ci">+{fmt(sum(s["interest"] for s in sav_stats))}원</div></div>'
            f'<div class="sc"><div class="l">총 만기 수령</div><div class="v ci">{fmt(sum(s["maturity_amount"] for s in sav_stats))}원</div></div>'
            f'<div class="sc"><div class="l">상품 수</div><div class="v cn">{len(sav_stats)}개</div></div>'
            '</div>'
        )

    def inv_row(i):
        gain = i['profit']
        gc = '#dc3545' if gain >= 0 else '#0d6efd'
        gs = '+' if gain >= 0 else ''
        gp = i['profit_pct']
        ticker_str = f' ({i["ticker"]})' if i['ticker'] else ''
        fx = i.get('exchange_rate') or None
        is_usd = i['itype'] == '해외주식' and fx
        avg_krw = int(i['avg_price'] * fx) if is_usd else i['avg_price']
        cur_krw = int(i['current_price'] * fx) if is_usd else i['current_price']
        usd_note = lambda p: f'<br><span style="font-size:10px;color:#aaa">${p:.2f}</span>' if is_usd else ''
        qty_unit = '개' if i['itype'] == '코인' else '주'
        return (
            '<tr>'
            f'<td><span class="badge bj">{i["itype"]}</span></td>'
            f'<td>{i["name"]}{ticker_str}</td>'
            f'<td style="text-align:right">{i["quantity"]:g}{qty_unit}</td>'
            f'<td style="text-align:right">{fmt(avg_krw)}원{usd_note(i["avg_price"])}</td>'
            f'<td style="text-align:right">{fmt(cur_krw)}원{usd_note(i["current_price"])}</td>'
            f'<td style="text-align:right;font-weight:700">{fmt(i["current_value"])}원</td>'
            f'<td style="text-align:right;color:{gc}">{gs}{fmt(gain)}원<br><span style="font-size:11px">({gs}{gp:.1f}%)</span></td>'
            '</tr>'
        )

    inv_gc = '#dc3545' if inv_gain_total >= 0 else '#0d6efd'
    inv_gs = '+' if inv_gain_total >= 0 else ''
    inv_rc = '#dc3545' if inv_return_rate >= 0 else '#0d6efd'
    inv_rs = '+' if inv_return_rate >= 0 else ''
    if investments:
        inv_html = (
            '<table><thead><tr><th>유형</th><th>종목</th><th style="text-align:right">수량</th>'
            '<th style="text-align:right">평균단가</th><th style="text-align:right">현재가</th>'
            '<th style="text-align:right">평가금액</th><th style="text-align:right">손익</th></tr></thead>'
            '<tbody>' + ''.join(inv_row(i) for i in investments) + '</tbody></table>'
            '<div class="sg" style="margin-top:12px">'
            f'<div class="sc"><div class="l">총 평가금액</div><div class="v cn">{fmt(inv_total)}원</div></div>'
            f'<div class="sc"><div class="l">총 손익</div><div class="v" style="color:{inv_gc}">{inv_gs}{fmt(inv_gain_total)}원</div></div>'
            f'<div class="sc"><div class="l">수익률</div><div class="v" style="color:{inv_rc}">{inv_rs}{inv_return_rate:.1f}%</div></div>'
            f'<div class="sc"><div class="l">종목 수</div><div class="v cn">{len(investments)}개</div></div>'
            '</div>'
        )
    else:
        inv_html = '<div class="empty">등록된 투자 종목이 없습니다</div>'

    def tx_row(tx):
        color = '#198754' if tx.type == 'income' else '#dc3545'
        sign = '+' if tx.type == 'income' else '-'
        badge_cls = 'bi' if tx.type == 'income' else 'be'
        label = '수입' if tx.type == 'income' else '지출'
        return (
            '<tr>'
            f'<td>{tx.date}</td>'
            f'<td><span class="badge {badge_cls}">{label}</span></td>'
            f'<td>{tx.category or "—"}</td>'
            f'<td>{tx.description or "—"}</td>'
            f'<td>{tx.card or "—"}</td>'
            f'<td style="text-align:right;font-weight:600;color:{color}">{sign}{fmt(tx.amount)}원</td>'
            '</tr>'
        )

    display_txs = all_txs[:50]
    txs_html = (
        '<table><thead><tr><th>날짜</th><th>유형</th><th>카테고리</th><th>설명</th><th>카드/계좌</th>'
        '<th style="text-align:right">금액</th></tr></thead>'
        '<tbody>' + ''.join(tx_row(tx) for tx in display_txs) + '</tbody></table>'
    ) if display_txs else '<div class="empty">거래 내역이 없습니다</div>'

    css = """*{box-sizing:border-box;margin:0;padding:0;-webkit-print-color-adjust:exact;print-color-adjust:exact}
body{font-family:'Malgun Gothic','맑은 고딕','Apple SD Gothic Neo',sans-serif;color:#222;background:#fff;padding:40px;font-size:13px;line-height:1.5}
.hdr{margin-bottom:32px;padding-bottom:20px;border-bottom:3px solid #b088f9}
.hdr h1{font-size:22px;font-weight:700;color:#b088f9;margin-bottom:4px}
.hdr .meta{font-size:11px;color:#888}
h2{font-size:14px;font-weight:700;color:#7c4fbf;background:#f0eaff;padding:12px 20px;margin:0;border-bottom:1.5px solid #e0d0fd}
.sec{margin-bottom:20px;border:1.5px solid #e0d0fd;border-radius:14px;overflow:hidden}
.si{padding:20px}
.sg{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:6px}
.sc{border:1px solid #e8e8e8;border-radius:10px;padding:14px 12px;text-align:center}
.sc .l{font-size:10px;color:#555;margin-bottom:5px}
.sc .v{font-size:15px;font-weight:700}
.ci{color:#198754}.ce{color:#dc3545}.cb{color:#b088f9}.cn{color:#333}
table{width:100%;border-collapse:collapse;font-size:12px;margin-bottom:4px}
thead th{background:#f8f5ff;color:#555;font-weight:700;padding:8px 10px;text-align:left;border-bottom:2px solid #e8d5ff}
tbody td{padding:8px 10px;border-bottom:1px solid #f5f5f5;vertical-align:middle}
tbody tr:last-child td{border-bottom:none}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:600}
.be{background:#ffe0e0;color:#dc3545}.bi{background:#d4edda;color:#198754}
.by{background:#f0e8fd;color:#b088f9}.bj{background:#f0e8fd;color:#b088f9}
.cp{border:1px solid #e8e8e8;border-radius:12px;padding:16px;margin-bottom:10px}
.cn2{font-size:14px;font-weight:700;margin-bottom:10px}
.ig{display:grid;grid-template-columns:repeat(3,1fr);gap:1px;background:#eee;border-radius:8px;overflow:hidden;margin-bottom:12px}
.ic{background:white;padding:8px 10px;text-align:center}
.ic .l{font-size:10px;color:#555;margin-bottom:3px}
.ic .v{font-size:12px;font-weight:600}
.pb{background:#f0f0f0;border-radius:6px;height:8px;overflow:hidden}
.pf{height:8px;border-radius:6px}
.pl{display:flex;justify-content:space-between;font-size:10px;color:#aaa;margin-top:3px}
.sp{border:1px solid #e8e8e8;border-radius:12px;padding:16px;margin-bottom:10px}
.sh{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px}
.sn{font-size:14px;font-weight:700}
.dd{font-size:12px;font-weight:700;color:#b088f9}
.sg2{display:grid;grid-template-columns:repeat(3,1fr);gap:1px;background:#eee;border-radius:8px;overflow:hidden;margin-bottom:10px}
.empty{color:#aaa;text-align:center;padding:16px;font-size:12px}
.footer{margin-top:40px;font-size:10px;color:#bbb;text-align:center;border-top:1px solid #eee;padding-top:16px}
.pdfbtn{position:fixed;top:16px;right:16px;z-index:9999;background:linear-gradient(135deg,#b088f9,#7baff0);color:white;border:none;border-radius:14px;padding:12px 22px;font-size:15px;font-weight:700;cursor:pointer;box-shadow:0 4px 16px rgba(176,136,249,0.4)}
@media print{body{padding:20px}.pdfbtn{display:none}}
@media(max-width:600px){body{padding:16px}.sg{grid-template-columns:repeat(2,1fr)}table{font-size:11px}thead th,tbody td{padding:6px 6px}}"""

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>포트폴리오 — {name_display}</title>
<style>{css}</style>
</head>
<body>
<button class="pdfbtn" onclick="window.print()">⬇ PDF 저장</button>
<div class="hdr">
  <h1>재무 포트폴리오</h1>
  <div class="meta">계정: {name_display} &nbsp;|&nbsp; 출력일: {date_str}</div>
</div>

<div class="sec">
  <h2>순자산 요약</h2>
  <div class="si"><div class="sg">
    <div class="sc"><div class="l">순자산</div><div class="v cn">{fmt(net_worth)}원</div></div>
    <div class="sc"><div class="l">이달 수입</div><div class="v ci">{fmt(income_mo)}원</div></div>
    <div class="sc"><div class="l">이달 지출</div><div class="v ce">{fmt(expense_mo)}원</div></div>
    <div class="sc"><div class="l">이달 잔액</div><div class="v cb">{fmt(income_mo - expense_mo)}원</div></div>
  </div></div>
</div>

<div class="sec">
  <h2>자산 구성</h2>
  <div class="si">
    <div style="display:flex;flex-direction:column;align-items:center;gap:16px">
      {donut_svg}
      <div style="width:100%">{legend_html}</div>
    </div>
  </div>
</div>

<div class="sec">
  <h2>총 자산 추이 (최근 6개월)</h2>
  <div class="si">
    {bar_chart}
    <table>
      <thead><tr><th>월</th><th style="text-align:right">총 자산</th><th style="text-align:right">전월 대비</th></tr></thead>
      <tbody>{trend_rows}</tbody>
    </table>
  </div>
</div>

<div class="sec">
  <h2>카드 / 계좌 ({len(card_stats)}개)</h2>
  <div class="si">{cards_html}</div>
</div>

<div class="sec">
  <h2>예적금 ({len(sav_stats)}개)</h2>
  <div class="si">{savings_html}{sav_summary_html}</div>
</div>

<div class="sec">
  <h2>투자 ({len(investments)}개 종목)</h2>
  <div class="si">{inv_html}</div>
</div>

<div class="sec">
  <h2>거래 내역 (최근 {len(display_txs)}건 / 전체 {len(all_txs)}건)</h2>
  <div class="si">{txs_html}</div>
</div>

<div class="footer">생성: 나의 가계부 앱 &nbsp;|&nbsp; {name_display} &nbsp;|&nbsp; {date_str}</div>
<script>window.onload = function() {{ window.print(); }}</script>
</body>
</html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}

@app.route('/api/budget', methods=['GET', 'POST'])
@login_required
def api_budget():
    uid = session['user_id']
    current_month = datetime.now().strftime('%Y-%m')
    if request.method == 'POST':
        data = request.json or {}
        amount = int(data.get('amount', 0))
        existing = Budget.query.filter_by(month=current_month, user_id=uid).first()
        if existing:
            existing.amount = amount
        else:
            db.session.add(Budget(month=current_month, amount=amount, user_id=uid))
        db.session.commit()
        return jsonify({'ok': True})
    budget = Budget.query.filter_by(month=current_month, user_id=uid).first()
    all_txs = Transaction.query.filter_by(user_id=uid).all()
    expense_total = sum(tx.amount for tx in all_txs if tx.type == 'expense' and tx.date.startswith(current_month))

    cards = Card.query.filter_by(user_id=uid).all()
    excl_cats_budget = {c.name for c in Category.query.filter_by(user_id=uid, exclude_perf=True).all()}
    card_by_id = {c.id: c for c in cards}
    # linked_names[account_id] = [card_name, ...]
    linked_names = {}
    for c in cards:
        if c.linked_account_id:
            linked_names.setdefault(c.linked_account_id, []).append(c.name)

    card_stats = []
    for card in cards:
        initial_balance = card.account_balance or 0
        is_loan = initial_balance < 0
        linked_account_id = card.linked_account_id
        if is_loan:
            card_stats.append({
                'id': card.id, 'name': card.name,
                'initial_balance': initial_balance, 'total_income': 0,
                'total_expense': 0, 'balance': initial_balance,
                'spent': 0, 'target': card.monthly_target or 0, 'percent': 0,
                'tier1': card.tier1 or 20, 'tier2': card.tier2 or 50, 'tier3': card.tier3 or 80,
                'url': card.url or '', 'is_loan': True, 'linked_account_id': linked_account_id,
            })
        else:
            card_txs = [tx for tx in all_txs if tx.card == card.name]
            month_income = sum(tx.amount for tx in card_txs if tx.type == 'income' and tx.date.startswith(current_month))
            month_expense = sum(tx.amount for tx in card_txs if tx.type == 'expense' and tx.date.startswith(current_month))
            perf_spent = sum(tx.amount for tx in card_txs
                             if tx.type == 'expense' and tx.date.startswith(current_month)
                             and _is_perf_tx(tx, excl_cats_budget))
            # account card: balance includes all linked cards' transactions
            if card.id in linked_names:
                for lname in linked_names[card.id]:
                    ltxs = [tx for tx in all_txs if tx.card == lname]
                    month_income += sum(tx.amount for tx in ltxs if tx.type == 'income' and tx.date.startswith(current_month))
                    month_expense += sum(tx.amount for tx in ltxs if tx.type == 'expense' and tx.date.startswith(current_month))
            # linked card: show account card's balance
            if linked_account_id and linked_account_id in card_by_id:
                acc = card_by_id[linked_account_id]
                acc_initial = acc.account_balance or 0
                acc_txs_names = [acc.name] + linked_names.get(linked_account_id, [])
                acc_inc = sum(tx.amount for tx in all_txs if tx.card in acc_txs_names and tx.type == 'income' and tx.date.startswith(current_month))
                acc_exp = sum(tx.amount for tx in all_txs if tx.card in acc_txs_names and tx.type == 'expense' and tx.date.startswith(current_month))
                balance = acc_initial + acc_inc - acc_exp
            else:
                balance = initial_balance + month_income - month_expense
            percent = min(int(perf_spent / card.monthly_target * 100), 100) if card.monthly_target > 0 else 0
            card_stats.append({
                'id': card.id, 'name': card.name,
                'initial_balance': initial_balance, 'total_income': month_income,
                'total_expense': month_expense, 'balance': balance,
                'spent': perf_spent, 'target': card.monthly_target, 'percent': percent,
                'tier1': card.tier1 or 20, 'tier2': card.tier2 or 50, 'tier3': card.tier3 or 80,
                'url': card.url or '', 'is_loan': False, 'linked_account_id': linked_account_id,
            })

    savings = Savings.query.filter_by(user_id=uid).all()
    investments = Investment.query.filter_by(user_id=uid).all()
    _auto_fetch_investment_prices(investments)
    extra_deposits = {}
    for dep in SavingsDeposit.query.filter_by(user_id=uid).all():
        extra_deposits[dep.savings_id] = extra_deposits.get(dep.savings_id, 0) + dep.amount
    return jsonify({
        'budget_amount': budget.amount if budget else 0,
        'expense_total': expense_total,
        'current_month': current_month,
        'card_stats': card_stats,
        'savings': [_savings_stats(s, extra_deposits.get(s.id, 0)) for s in savings],
        'investments': [_investment_stats(i) for i in investments],
    })

@app.route('/api/savings', methods=['GET', 'POST'])
@login_required
def api_savings():
    uid = session['user_id']
    if request.method == 'POST':
        data = request.json or {}
        nd = data.get('notify_day')
        atd = data.get('auto_tx_day')
        db.session.add(Savings(
            user_id=uid,
            stype=data.get('stype', '예금'),
            bank=data.get('bank', ''),
            name=data['name'],
            amount=int(data.get('amount', 0)),
            interest_rate=float(data.get('interest_rate', 0)),
            interest_type=data.get('interest_type', '단리'),
            tax_type=data.get('tax_type', '일반과세'),
            start_date=data['start_date'],
            end_date=data['end_date'],
            notify_day=int(nd) if nd else None,
            auto_tx=bool(data.get('auto_tx', False)),
            auto_tx_day=int(atd) if atd else None,
            auto_tx_card=data.get('auto_tx_card', '') or '',
            bonus_amount=int(data['bonus_amount']) if data.get('bonus_amount') else None,
        ))
        db.session.commit()
        return jsonify({'ok': True})
    items = Savings.query.filter_by(user_id=uid).all()
    extra_deposits_sav = {}
    for dep in SavingsDeposit.query.filter_by(user_id=uid).all():
        extra_deposits_sav[dep.savings_id] = extra_deposits_sav.get(dep.savings_id, 0) + dep.amount
    return jsonify({'savings': [_savings_stats(s, extra_deposits_sav.get(s.id, 0)) for s in items]})

@app.route('/api/savings/<int:sid>', methods=['PUT', 'DELETE'])
@login_required
def api_saving(sid):
    uid = session['user_id']
    s = Savings.query.filter_by(id=sid, user_id=uid).first_or_404()
    if request.method == 'DELETE':
        db.session.delete(s)
        db.session.commit()
        return jsonify({'ok': True})
    data = request.json or {}
    s.stype = data.get('stype', s.stype)
    s.bank = data.get('bank', s.bank)
    s.name = data.get('name', s.name)
    s.amount = int(data.get('amount', s.amount))
    s.interest_rate = float(data.get('interest_rate', s.interest_rate))
    s.interest_type = data.get('interest_type', getattr(s, 'interest_type', '단리') or '단리')
    s.tax_type = data.get('tax_type', getattr(s, 'tax_type', '일반과세') or '일반과세')
    s.start_date = data.get('start_date', s.start_date)
    s.end_date = data.get('end_date', s.end_date)
    if 'notify_day' in data:
        nd = data['notify_day']
        s.notify_day = int(nd) if nd else None
    if 'auto_tx' in data:
        s.auto_tx = bool(data['auto_tx'])
    if 'auto_tx_day' in data:
        atd = data['auto_tx_day']
        s.auto_tx_day = int(atd) if atd else None
    if 'auto_tx_card' in data:
        s.auto_tx_card = data['auto_tx_card'] or ''
    if 'manual_count' in data:
        mc = data['manual_count']
        s.manual_count = int(mc) if mc is not None else None
    if 'is_paused' in data:
        s.is_paused = bool(data['is_paused'])
    if 'bonus_amount' in data:
        ba = data['bonus_amount']
        s.bonus_amount = int(ba) if ba else None
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/usd-rate')
@login_required
def get_usd_rate():
    try:
        import FinanceDataReader as fdr
        from datetime import date, timedelta
        start = (date.today() - timedelta(days=7)).strftime('%Y-%m-%d')
        fx = fdr.DataReader('USD/KRW', start)
        rate = int(float(fx['Close'].iloc[-1])) if not fx.empty else 1380
        return jsonify({'ok': True, 'rate': rate})
    except Exception:
        return jsonify({'ok': True, 'rate': 1380})

@app.route('/api/investments', methods=['GET', 'POST'])
@login_required
def api_investments():
    uid = session['user_id']
    if request.method == 'POST':
        data = request.json or {}
        inv = Investment(
            user_id=uid,
            itype=data.get('itype', '국내주식'),
            name=data.get('name', ''),
            ticker=data.get('ticker', ''),
            quantity=float(data.get('quantity', 0)),
            avg_price=float(data.get('avg_price', 0)),
            current_price=float(data['current_price']) if data.get('current_price') not in (None, '') else None,
            exchange_rate=float(data['exchange_rate']) if data.get('exchange_rate') not in (None, '') else None,
            memo=data.get('memo', ''),
            account_type=data.get('account_type', '일반'),
        )
        db.session.add(inv)
        db.session.commit()
        return jsonify({'ok': True})
    items = Investment.query.filter_by(user_id=uid).all()
    return jsonify({'investments': [_investment_stats(i) for i in items]})

@app.route('/api/investments/<int:iid>', methods=['PUT', 'DELETE'])
@login_required
def api_investment(iid):
    uid = session['user_id']
    inv = Investment.query.filter_by(id=iid, user_id=uid).first_or_404()
    if request.method == 'DELETE':
        db.session.delete(inv)
        db.session.commit()
        return jsonify({'ok': True})
    data = request.json or {}
    inv.itype = data.get('itype', inv.itype)
    inv.name = data.get('name', inv.name)
    inv.ticker = data.get('ticker', inv.ticker)
    inv.quantity = float(data.get('quantity', inv.quantity))
    inv.avg_price = float(data.get('avg_price', inv.avg_price))
    inv.current_price = float(data['current_price']) if data.get('current_price') not in (None, '') else None
    inv.exchange_rate = float(data['exchange_rate']) if data.get('exchange_rate') not in (None, '') else inv.exchange_rate
    inv.memo = data.get('memo', inv.memo)
    if 'account_type' in data:
        inv.account_type = data['account_type']
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/investments/price')
@login_required
def fetch_investment_price():
    from datetime import date, timedelta
    import FinanceDataReader as fdr
    ticker = request.args.get('ticker', '').strip()
    itype = request.args.get('itype', '')
    if not ticker:
        return jsonify({'ok': False, 'error': '티커를 입력하세요'}), 400
    try:
        if itype == '코인' or ticker.upper().startswith('KRW-'):
            import urllib.request as _ur, json as _js
            url = f'https://api.upbit.com/v1/ticker?markets={ticker.upper()}'
            req = _ur.Request(url, headers={'User-Agent': 'Mozilla/5.0', 'Accept': 'application/json'})
            with _ur.urlopen(req, timeout=6) as r:
                d = _js.loads(r.read())
            price = d[0]['trade_price']
            return jsonify({'ok': True, 'price': price, 'price_krw': price, 'currency': 'KRW'})
        start = (date.today() - timedelta(days=7)).strftime('%Y-%m-%d')
        is_kr = itype == '국내주식' or (itype == 'ETF' and ticker.replace('.', '').isdigit())
        if is_kr:
            t = ticker.replace('.KS', '').replace('.KQ', '')
            df = fdr.DataReader(t, start)
            if df.empty:
                return jsonify({'ok': False, 'error': '데이터를 가져올 수 없습니다'}), 400
            price = float(df['Close'].iloc[-1])
            return jsonify({'ok': True, 'price': price, 'price_krw': int(price), 'currency': 'KRW'})
        else:
            df = fdr.DataReader(ticker, start)
            if df.empty:
                return jsonify({'ok': False, 'error': '데이터를 가져올 수 없습니다'}), 400
            price = float(df['Close'].iloc[-1])
            try:
                fx = fdr.DataReader('USD/KRW', start)
                usd_krw = float(fx['Close'].iloc[-1]) if not fx.empty else 1380
            except Exception:
                usd_krw = 1380
            return jsonify({'ok': True, 'price': price, 'price_krw': int(price * usd_krw), 'currency': 'USD'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


ADMIN_EMAIL = 'song57290@gmail.com'

@app.route('/api/notices', methods=['GET', 'POST'])
@login_required
def api_notices():
    uid = session['user_id']
    u = User.query.get(uid)
    is_admin = u and u.email == ADMIN_EMAIL
    if request.method == 'POST':
        if not is_admin:
            return jsonify({'ok': False, 'error': '권한이 없습니다'}), 403
        data = request.json or {}
        title = (data.get('title') or '').strip()
        content = (data.get('content') or '').strip()
        if not title or not content:
            return jsonify({'ok': False, 'error': '제목과 내용을 입력하세요'}), 400
        db.session.add(Notice(user_id=uid, title=title, content=content, created_at=datetime.now()))
        db.session.commit()
        return jsonify({'ok': True})
    notices = Notice.query.order_by(Notice.created_at.desc()).all()
    return jsonify([{
        'id': n.id, 'title': n.title, 'content': n.content,
        'created_at': n.created_at.strftime('%Y.%m.%d'),
        'is_admin': is_admin,
    } for n in notices])

@app.route('/api/notices/<int:nid>', methods=['DELETE', 'PUT'])
@login_required
def api_notice(nid):
    uid = session['user_id']
    u = User.query.get(uid)
    if not u or u.email != ADMIN_EMAIL:
        return jsonify({'ok': False, 'error': '권한이 없습니다'}), 403
    n = Notice.query.get_or_404(nid)
    if request.method == 'DELETE':
        db.session.delete(n)
        db.session.commit()
        return jsonify({'ok': True})
    data = request.json or {}
    title = (data.get('title') or '').strip()
    content = (data.get('content') or '').strip()
    if not title or not content:
        return jsonify({'ok': False, 'error': '제목과 내용을 입력하세요'}), 400
    n.title = title
    n.content = content
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/salary', methods=['GET', 'PUT'])
@login_required
def api_salary():
    uid = session['user_id']
    if request.method == 'GET':
        cfg = SalaryConfig.query.filter_by(user_id=uid).first()
        allocs = BudgetAllocation.query.filter_by(user_id=uid).all()
        fixed = FixedExpense.query.filter_by(user_id=uid).all()
        current_month = datetime.now().strftime('%Y-%m')
        txs = Transaction.query.filter_by(user_id=uid).filter(Transaction.date.like(f'{current_month}%')).all()
        actual = {}
        for tx in txs:
            if tx.type == 'expense':
                actual[tx.category] = actual.get(tx.category, 0) + tx.amount
        fixed_list = [{'id': f.id, 'name': f.name, 'amount': f.amount, 'day_of_month': f.day_of_month,
                        'category': f.category, 'auto_register': bool(f.auto_register),
                        'tx_type': f.tx_type or 'expense', 'tx_card': f.tx_card or '',
                        'item_type': 'fixed'} for f in fixed]
        savings_all = Savings.query.filter_by(user_id=uid).all()
        for s in savings_all:
            if s.stype not in ('적금', '청약'):
                continue
            auto_tx = bool(getattr(s, 'auto_tx', False))
            if not auto_tx or bool(getattr(s, 'is_paused', False)):
                continue
            fixed_list.append({'id': s.id, 'name': s.name, 'amount': s.amount,
                                'day_of_month': getattr(s, 'auto_tx_day', None),
                                'category': '저축', 'auto_register': auto_tx,
                                'tx_type': 'expense', 'tx_card': getattr(s, 'auto_tx_card', '') or '',
                                'item_type': 'savings', 'stype': s.stype})
        return jsonify({
            'salary': {'amount': cfg.amount if cfg else 0, 'pay_day': cfg.pay_day if cfg else None},
            'allocations': [{'id': a.id, 'category_name': a.category_name, 'percent': a.percent} for a in allocs],
            'fixed_expenses': fixed_list,
            'actual': actual,
        })
    data = request.get_json()
    cfg = SalaryConfig.query.filter_by(user_id=uid).first()
    if cfg:
        cfg.amount = data.get('amount', cfg.amount)
        cfg.pay_day = data.get('pay_day', cfg.pay_day)
    else:
        db.session.add(SalaryConfig(user_id=uid, amount=data.get('amount', 0), pay_day=data.get('pay_day')))
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/salary/allocations', methods=['PUT'])
@login_required
def api_salary_allocations():
    uid = session['user_id']
    data = request.get_json()
    BudgetAllocation.query.filter_by(user_id=uid).delete()
    for a in data:
        if a.get('percent', 0) > 0:
            db.session.add(BudgetAllocation(user_id=uid, category_name=a['category_name'], percent=a['percent']))
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/salary/fixed', methods=['POST'])
@login_required
def api_fixed_expenses_add():
    uid = session['user_id']
    data = request.get_json()
    f = FixedExpense(user_id=uid, name=data['name'], amount=data['amount'],
                     day_of_month=data.get('day_of_month'), category=data.get('category', ''),
                     auto_register=bool(data.get('auto_register', False)),
                     tx_type=data.get('tx_type', 'expense'), tx_card=data.get('tx_card', ''))
    db.session.add(f)
    db.session.commit()
    return jsonify({'id': f.id, 'name': f.name, 'amount': f.amount, 'day_of_month': f.day_of_month, 'category': f.category, 'auto_register': bool(f.auto_register), 'tx_type': f.tx_type, 'tx_card': f.tx_card})

@app.route('/api/salary/fixed/<int:fid>', methods=['PUT', 'DELETE'])
@login_required
def api_fixed_expense(fid):
    uid = session['user_id']
    f = FixedExpense.query.filter_by(id=fid, user_id=uid).first_or_404()
    if request.method == 'DELETE':
        db.session.delete(f)
        db.session.commit()
        return jsonify({'ok': True})
    data = request.get_json()
    f.name = data.get('name', f.name)
    f.amount = data.get('amount', f.amount)
    f.day_of_month = data.get('day_of_month', f.day_of_month)
    f.category = data.get('category', f.category)
    f.auto_register = bool(data.get('auto_register', f.auto_register))
    f.tx_type = data.get('tx_type', f.tx_type or 'expense')
    f.tx_card = data.get('tx_card', f.tx_card or '')
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/savings/<int:sid>/deposits', methods=['GET', 'POST'])
@login_required
def api_savings_deposits(sid):
    uid = session['user_id']
    Savings.query.filter_by(id=sid, user_id=uid).first_or_404()
    if request.method == 'POST':
        data = request.get_json()
        dep = SavingsDeposit(savings_id=sid, user_id=uid,
                             amount=int(data.get('amount', 0)),
                             date=data.get('date', datetime.now().strftime('%Y-%m-%d')),
                             memo=data.get('memo', ''))
        db.session.add(dep)
        db.session.commit()
        return jsonify({'id': dep.id, 'amount': dep.amount, 'date': dep.date, 'memo': dep.memo})
    deps = SavingsDeposit.query.filter_by(savings_id=sid, user_id=uid).order_by(SavingsDeposit.date.desc()).all()
    return jsonify([{'id': d.id, 'amount': d.amount, 'date': d.date, 'memo': d.memo} for d in deps])

@app.route('/api/savings/deposits/<int:did>', methods=['DELETE'])
@login_required
def api_savings_deposit_delete(did):
    uid = session['user_id']
    dep = SavingsDeposit.query.filter_by(id=did, user_id=uid).first_or_404()
    db.session.delete(dep)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/pending-registers', methods=['GET'])
@login_required
def api_pending_registers():
    uid = session['user_id']
    today = datetime.now()
    today_day = today.day
    current_month = today.strftime('%Y-%m')
    pending = []
    # 고정 지출
    for f in FixedExpense.query.filter_by(user_id=uid, auto_register=True).all():
        if f.day_of_month != today_day:
            continue
        already = Transaction.query.filter_by(user_id=uid).filter(
            Transaction.date.like(f'{current_month}%'),
            Transaction.description == f'[자동]{f.name}',
        ).first()
        if not already:
            pending.append({'item_type': 'fixed', 'id': f.id, 'name': f.name, 'amount': f.amount,
                            'category': f.category or '', 'tx_type': f.tx_type or 'expense',
                            'tx_card': f.tx_card or ''})
    # 적금/청약 자동이체
    for s in Savings.query.filter_by(user_id=uid, auto_tx=True).all():
        if s.stype not in ('적금', '청약'):
            continue
        if bool(getattr(s, 'is_paused', False)):
            continue
        atd = getattr(s, 'auto_tx_day', None)
        if not atd or atd != today_day:
            continue
        desc = f'[자동이체]{s.name}'
        already = Transaction.query.filter_by(user_id=uid).filter(
            Transaction.date.like(f'{current_month}%'),
            Transaction.description == desc,
        ).first()
        if not already:
            pending.append({'item_type': 'savings', 'id': s.id, 'name': s.name, 'amount': s.amount,
                            'category': '저축', 'tx_type': 'expense',
                            'tx_card': getattr(s, 'auto_tx_card', '') or ''})
    return jsonify(pending)

@app.route('/api/salary/fixed/<int:fid>/register', methods=['POST'])
@login_required
def api_fixed_register(fid):
    uid = session['user_id']
    f = FixedExpense.query.filter_by(id=fid, user_id=uid).first_or_404()
    today = datetime.now().strftime('%Y-%m-%d')
    tx = Transaction(date=today, type=f.tx_type or 'expense',
                     category=f.category or '기타', description=f'[자동]{f.name}',
                     amount=f.amount, card=f.tx_card or None, user_id=uid)
    db.session.add(tx)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/savings/<int:sid>/auto-register', methods=['POST'])
@login_required
def api_savings_auto_register(sid):
    uid = session['user_id']
    s = Savings.query.filter_by(id=sid, user_id=uid).first_or_404()
    today = datetime.now().strftime('%Y-%m-%d')
    tx = Transaction(date=today, type='expense',
                     category='저축', description=f'[자동이체]{s.name}',
                     amount=s.amount, card=getattr(s, 'auto_tx_card', '') or None, user_id=uid)
    db.session.add(tx)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/help', methods=['GET'])
def api_help_list():
    items = HelpItem.query.order_by(HelpItem.position).all()
    return jsonify([{'id': h.id, 'icon': h.icon, 'title': h.title, 'desc': h.desc} for h in items])

@app.route('/api/help/<int:item_id>', methods=['PUT'])
@login_required
def api_help_item(item_id):
    u = User.query.get(session['user_id'])
    if u.email != ADMIN_EMAIL:
        return jsonify({'error': 'Forbidden'}), 403
    h = HelpItem.query.get_or_404(item_id)
    data = request.get_json()
    if 'icon' in data: h.icon = data['icon']
    if 'title' in data: h.title = data['title']
    if 'desc' in data: h.desc = data['desc']
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/update-notice-config', methods=['GET'])
def api_update_notice_config_get():
    config = AppConfig.query.get('update_notice')
    if config:
        return jsonify(json.loads(config.value))
    return jsonify({'version': '', 'date': '', 'updates': []})

@app.route('/api/update-notice-config', methods=['PUT'])
@login_required
def api_update_notice_config_put():
    u = User.query.get(session['user_id'])
    if u.email != ADMIN_EMAIL:
        return jsonify({'error': 'Forbidden'}), 403
    data = request.get_json()
    config = AppConfig.query.get('update_notice')
    if config:
        config.value = json.dumps(data, ensure_ascii=False)
    else:
        db.session.add(AppConfig(key='update_notice', value=json.dumps(data, ensure_ascii=False)))
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/portfolio')
@login_required
def api_portfolio():
    uid = session['user_id']
    user = User.query.get(uid)
    current_month = datetime.now().strftime('%Y-%m')
    transactions = Transaction.query.filter_by(user_id=uid).order_by(Transaction.date.desc()).all()
    month_txs = [tx for tx in transactions if tx.date.startswith(current_month)]
    excl_stat_cats_port = {c.name for c in Category.query.filter_by(user_id=uid, exclude_stats=True).all()}
    income_total = sum(tx.amount for tx in month_txs if tx.type == 'income' and _is_stats_tx(tx, excl_stat_cats_port))
    expense_total = sum(tx.amount for tx in month_txs if tx.type == 'expense' and _is_stats_tx(tx, excl_stat_cats_port))
    cards = Card.query.filter_by(user_id=uid).all()
    card_stats = []
    for card in cards:
        card_txs = [tx for tx in transactions if tx.card == card.name]
        current_balance = card.account_balance or 0
        month_income = sum(tx.amount for tx in card_txs if tx.type == 'income' and tx.date.startswith(current_month))
        spent = sum(tx.amount for tx in card_txs
                    if tx.type == 'expense' and tx.date.startswith(current_month)
                    and _is_stats_tx(tx, excl_stat_cats_port))
        initial_balance = current_balance - month_income + spent
        percent = min(int(spent / card.monthly_target * 100), 100) if card.monthly_target > 0 else 0
        card_stats.append({'name': card.name, 'initial_balance': initial_balance,
                           'balance': current_balance,
                           'month_income': month_income, 'spent': spent,
                           'target': card.monthly_target, 'percent': percent})
    savings_list = Savings.query.filter_by(user_id=uid).all()
    extra_deposits_port = {}
    for dep in SavingsDeposit.query.filter_by(user_id=uid).all():
        extra_deposits_port[dep.savings_id] = extra_deposits_port.get(dep.savings_id, 0) + dep.amount
    savings_stats = [_savings_stats(s, extra_deposits_port.get(s.id, 0)) for s in savings_list]
    inv_list = Investment.query.filter_by(user_id=uid).all()
    _auto_fetch_investment_prices(inv_list)
    budget = Budget.query.filter_by(month=current_month, user_id=uid).first()
    investments = [_investment_stats(i) for i in inv_list]
    net_worth = sum(c['balance'] for c in card_stats) + sum(s['current_paid'] for s in savings_stats) + sum(i['current_value'] for i in investments)
    inv_total = sum(i['current_value'] for i in investments)
    inv_gain_total = sum(i['profit'] for i in investments)
    inv_cost_total = sum(i['purchase_value'] for i in investments)
    inv_return_rate = round(inv_gain_total / inv_cost_total * 100, 2) if inv_cost_total else 0
    return jsonify({
        'user': {'email': user.email, 'nickname': user.nickname},
        'current_month': current_month,
        'summary': {'income': income_total, 'expense': expense_total,
                    'balance': income_total - expense_total, 'tx_count': len(transactions)},
        'net_worth': net_worth,
        'cards': card_stats,
        'savings': savings_stats,
        'savings_summary': {
            'total_principal': sum(s['amount'] for s in savings_stats),
            'total_interest': sum(s['interest'] for s in savings_stats),
            'total_maturity': sum(s['maturity_amount'] for s in savings_stats),
        },
        'investments': investments,
        'investments_summary': {'total_value': inv_total, 'total_gain': inv_gain_total, 'count': len(investments), 'return_rate': inv_return_rate, 'total_cost': inv_cost_total},
        'budget': budget.amount if budget else 0,
        'transactions': [{'date': tx.date, 'type': tx.type, 'category': tx.category,
                          'description': tx.description or '', 'amount': tx.amount, 'card': tx.card or ''}
                         for tx in transactions],
    })

@app.route('/api/categories', methods=['GET', 'POST'])
@login_required
def api_categories():
    uid = session['user_id']
    if request.method == 'POST':
        data = request.json or {}
        name = data.get('name', '').strip()
        icon = data.get('icon', '📦').strip()
        cat_type = data.get('type', 'expense')
        if name and not Category.query.filter_by(name=name, user_id=uid).first():
            max_pos = db.session.query(db.func.max(Category.position)).filter(Category.user_id == uid).scalar() or 0
            db.session.add(Category(name=name, icon=icon, position=max_pos + 1, cat_type=cat_type, user_id=uid))
            db.session.commit()
        return jsonify({'ok': True})
    expense = Category.query.filter_by(cat_type='expense', user_id=uid).order_by(Category.position, Category.id).all()
    income = Category.query.filter_by(cat_type='income', user_id=uid).order_by(Category.position, Category.id).all()
    return jsonify({
        'expense': [{'id': c.id, 'name': c.name, 'icon': c.icon, 'type': c.cat_type, 'exclude_perf': bool(c.exclude_perf), 'exclude_stats': bool(c.exclude_stats)} for c in expense],
        'income': [{'id': c.id, 'name': c.name, 'icon': c.icon, 'type': c.cat_type, 'exclude_perf': bool(c.exclude_perf), 'exclude_stats': bool(c.exclude_stats)} for c in income],
    })

@app.route('/api/categories/reorder', methods=['POST'])
@login_required
def api_reorder_categories():
    uid = session['user_id']
    ids = (request.json or {}).get('ids', [])
    for i, cat_id in enumerate(ids):
        cat = Category.query.filter_by(id=cat_id, user_id=uid).first()
        if cat:
            cat.position = i
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/categories/<int:cat_id>', methods=['PUT', 'DELETE'])
@login_required
def api_category(cat_id):
    uid = session['user_id']
    cat = Category.query.filter_by(id=cat_id, user_id=uid).first_or_404()
    if request.method == 'DELETE':
        db.session.delete(cat)
        db.session.commit()
        return jsonify({'ok': True})
    data = request.json or {}
    cat.name = data.get('name', cat.name).strip()
    cat.icon = data.get('icon', cat.icon).strip()
    if 'exclude_perf' in data:
        cat.exclude_perf = bool(data['exclude_perf'])
    if 'exclude_stats' in data:
        cat.exclude_stats = bool(data['exclude_stats'])
    db.session.commit()
    return jsonify({'ok': True})

# ── SPA entry point ───────────────────────────────────────────────────────────

_DIST_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'frontend', 'dist')
_DIST_INDEX = os.path.join(_DIST_DIR, 'index.html')

@app.route('/assets/<path:filename>')
def serve_dist_assets(filename):
    return send_from_directory(os.path.join(_DIST_DIR, 'assets'), filename)

@app.route('/.well-known/assetlinks.json')
def assetlinks():
    from flask import Response
    data = [{
        'relation': ['delegate_permission/common.handle_all_urls'],
        'target': {
            'namespace': 'android_app',
            'package_name': 'app.gaegyebu',
            'sha256_cert_fingerprints': [
                '9A:52:C8:28:82:F8:A7:77:E1:CE:0A:93:7F:D2:8E:DB:09:37:C3:C4:09:AF:FF:BF:21:4C:39:9C:D8:2E:3C:76'
            ]
        }
    }]
    return Response(json.dumps(data), mimetype='application/json')

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_spa(path):
    if os.path.exists(_DIST_INDEX):
        return send_file(_DIST_INDEX)
    return 'Frontend not built. Run: cd frontend && npm run build', 503

# ── Excel import helpers ──────────────────────────────────────────────────────
def _parse_date(val):
    if val is None: return None
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    s = str(val).strip().split(' ')[0].split('T')[0]
    for fmt in ('%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d'):
        try: return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except: pass
    if len(s) == 8 and s.isdigit():
        return f'{s[:4]}-{s[4:6]}-{s[6:8]}'
    return None

def _parse_amount(val):
    if val is None: return None
    if isinstance(val, (int, float)): return int(abs(val)) if val != 0 else None
    s = str(val).replace(',', '').replace('원', '').replace(' ', '').strip()
    try: v = float(s); return int(abs(v)) if v != 0 else None
    except: return None

def _parse_type(val):
    if val is None: return None
    s = str(val).strip()
    for kw in ('출금', '지출', '결제', 'expense'):
        if kw in s: return 'expense'
    for kw in ('입금', '수입', '이자', 'income'):
        if kw in s: return 'income'
    return None

_DATE_H = {'날짜','거래일자','거래일','일자','거래날짜','날짜(time)'}
_TYPE_H = {'유형','구분','거래구분','거래유형','입출금구분','입출금'}
_DESC_H = {'내용','적요','거래내용','메모','설명','항목','가맹점명','사용처','적요내용','거래처'}
_AMT_H  = {'금액','거래금액','금액(원)','거래금액(원)'}
_DEB_H  = {'출금','출금액','지출금액','출금금액','출금(원)','출금금액(원)','출금액(원)'}
_CRD_H  = {'입금','입금액','수입금액','입금금액','입금(원)','입금금액(원)','입금액(원)'}
_CAT_H  = {'카테고리','분류'}
_CARD_H = {'카드','카드명','결제카드'}

def _detect_cols(header):
    cols = {}
    for i, h in enumerate(header):
        if h is None: continue
        h = str(h).strip().replace(' ', '')
        if h in _DATE_H: cols.setdefault('date', i)
        elif h in _TYPE_H: cols.setdefault('type', i)
        elif h in _DESC_H: cols.setdefault('desc', i)
        elif h in _AMT_H: cols.setdefault('amount', i)
        elif h in _DEB_H: cols['debit'] = i
        elif h in _CRD_H: cols['credit'] = i
        elif h in _CAT_H: cols.setdefault('category', i)
        elif h in _CARD_H: cols.setdefault('card', i)
    return cols

@app.route('/import/template')
def import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '내역'
    ws.append(['날짜', '유형', '카테고리', '설명', '금액', '카드'])
    ws.append(['2026-06-17', '지출', '식사', '스타벅스', 50000, '신한카드'])
    ws.append(['2026-06-17', '수입', '기타', '월급', 3000000, ''])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='가계부_양식.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/import', methods=['POST'])
def import_excel():
    uid = session.get('user_id')
    if not uid:
        return redirect('/login')
    file = request.files.get('file')
    fname = (file.filename or '').lower()
    if not file or not (fname.endswith('.xlsx') or fname.endswith('.xls')):
        return redirect('/?import_error=파일 형식 오류 (.xlsx 또는 .xls)')

    all_rows = []
    try:
        if fname.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row))
        else:
            wb = xlrd.open_workbook(file_contents=file.read())
            ws = wb.sheet_by_index(0)
            for i in range(ws.nrows):
                parsed = []
                for cell in ws.row(i):
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        parsed.append(xlrd.xldate_as_datetime(cell.value, wb.datemode).strftime('%Y-%m-%d'))
                    elif cell.ctype == xlrd.XL_CELL_EMPTY:
                        parsed.append(None)
                    else:
                        parsed.append(cell.value)
                all_rows.append(parsed)
    except Exception as e:
        app.logger.exception('Excel open error')
        from urllib.parse import quote
        return redirect('/?import_error=' + quote(str(e)[:120]))

    if not all_rows:
        return redirect('/?import_error=빈 파일입니다')

    def serialize(v):
        if v is None: return None
        if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
        return str(v)
    all_rows = [[serialize(c) for c in row] for row in all_rows]

    fd, path = tempfile.mkstemp(suffix='.json', prefix='impx_')
    with os.fdopen(fd, 'w', encoding='utf-8') as f:
        json.dump(all_rows, f)

    header_row = 0
    auto_cols = {}
    for ri, row in enumerate(all_rows[:15]):
        c = _detect_cols(row)
        if 'date' in c or 'debit' in c or 'credit' in c or 'amount' in c:
            header_row = ri
            auto_cols = c
            break

    from urllib.parse import urlencode
    params = urlencode({'tmp': os.path.basename(path), 'hr': header_row,
                        **{k: v for k, v in auto_cols.items()}})
    return redirect(url_for('import_map') + '?' + params)

@app.route('/import/map')
def import_map():
    uid = session.get('user_id')
    if not uid:
        return redirect('/login')
    tmp = request.args.get('tmp', '')
    if not tmp.startswith('impx_'):
        return redirect('/')
    path = os.path.join(tempfile.gettempdir(), tmp)
    if not os.path.exists(path):
        return redirect('/')
    with open(path, encoding='utf-8') as f:
        all_rows = json.load(f)
    header_row = int(request.args.get('hr', 0))
    headers = all_rows[header_row]
    preview = all_rows[header_row + 1 : header_row + 6]
    auto_cols = {k: int(v) for k, v in request.args.items()
                 if k not in ('tmp', 'hr') and v.lstrip('-').isdigit()}
    col_samples = []
    for ci in range(len(headers)):
        sample = ''
        for row in preview:
            if ci < len(row) and row[ci] is not None and str(row[ci]).strip():
                sample = str(row[ci]).strip()
                break
        col_samples.append(sample)

    return render_template('import_map.html', headers=headers, preview=preview,
                           auto_cols=auto_cols, tmp=tmp, header_row=header_row,
                           col_samples=col_samples)

@app.route('/import/confirm', methods=['POST'])
def import_confirm():
    uid = session.get('user_id')
    if not uid:
        return redirect('/login')
    tmp = request.form.get('tmp', '')
    if not tmp.startswith('impx_'):
        return redirect('/')
    path = os.path.join(tempfile.gettempdir(), tmp)
    if not os.path.exists(path):
        return redirect('/?import_error=세션이 만료되었습니다. 다시 업로드해주세요.')
    with open(path, encoding='utf-8') as f:
        all_rows = json.load(f)
    os.remove(path)

    header_row = int(request.form.get('header_row', 0))

    def gi(name):
        v = request.form.get(name, '')
        return int(v) if v.lstrip('-').isdigit() else -1

    col_date   = gi('col_date')
    col_debit  = gi('col_debit')
    col_credit = gi('col_credit')
    col_amount = gi('col_amount')
    col_type   = gi('col_type')
    col_desc   = gi('col_desc')
    col_cat    = gi('col_cat')

    parsed_rows = []
    skipped_rows = []
    row_num_offset = header_row + 2  # 1-based, header 다음 줄부터
    for ri, row in enumerate(all_rows[header_row + 1:]):
        row_num = ri + row_num_offset
        raw_preview = ', '.join(str(c) for c in row if c is not None and str(c).strip())[:80]
        if not any(row):
            continue
        try:
            date_val = _parse_date(row[col_date]) if 0 <= col_date < len(row) else None
            if not date_val:
                skipped_rows.append({'row': row_num, 'preview': raw_preview, 'reason': '날짜 인식 불가'})
                continue

            if col_debit >= 0 and col_credit >= 0:
                debit  = _parse_amount(row[col_debit])  if col_debit  < len(row) else None
                credit = _parse_amount(row[col_credit]) if col_credit < len(row) else None
                if debit:    tx_type, amount = 'expense', debit
                elif credit: tx_type, amount = 'income',  credit
                else:
                    skipped_rows.append({'row': row_num, 'preview': raw_preview, 'reason': '출금/입금 금액 없음'})
                    continue
            elif col_amount >= 0 and col_type >= 0:
                tx_type = _parse_type(row[col_type])
                amount  = _parse_amount(row[col_amount])
                if not tx_type or not amount:
                    skipped_rows.append({'row': row_num, 'preview': raw_preview, 'reason': '금액 또는 유형 인식 불가'})
                    continue
            elif col_amount >= 0:
                amount = _parse_amount(row[col_amount])
                if not amount:
                    skipped_rows.append({'row': row_num, 'preview': raw_preview, 'reason': '금액 인식 불가'})
                    continue
                tx_type = 'expense'
            else:
                skipped_rows.append({'row': row_num, 'preview': raw_preview, 'reason': '금액 컬럼 미지정'})
                continue

            desc_val = str(row[col_desc]).strip() if 0 <= col_desc < len(row) and row[col_desc] else ''
            cat_val  = str(row[col_cat]).strip()  if 0 <= col_cat  < len(row) and row[col_cat]  else ''

            parsed_rows.append({
                'date': date_val, 'type': tx_type,
                'description': desc_val, 'amount': amount, 'category': cat_val,
            })
        except Exception as e:
            skipped_rows.append({'row': row_num, 'preview': raw_preview, 'reason': f'파싱 오류: {str(e)[:40]}'})

    # 파싱된 내역을 임시파일에 저장 후 카테고리 선택 페이지로
    fd, cat_path = tempfile.mkstemp(suffix='.json', prefix='impc_')
    with os.fdopen(fd, 'w', encoding='utf-8') as f:
        json.dump({'rows': parsed_rows, 'skipped_rows': skipped_rows}, f)

    from urllib.parse import urlencode
    return redirect('/import/categorize?' + urlencode({'tmp': os.path.basename(cat_path)}))


@app.route('/import/categorize')
def import_categorize():
    uid = session.get('user_id')
    if not uid:
        return redirect('/login')
    tmp = request.args.get('tmp', '')
    if not tmp.startswith('impc_'):
        return redirect('/')
    path = os.path.join(tempfile.gettempdir(), tmp)
    if not os.path.exists(path):
        return redirect('/?import_error=세션이 만료되었습니다. 다시 업로드해주세요.')
    with open(path, encoding='utf-8') as f:
        data = json.load(f)

    cats_expense = Category.query.filter(
        (Category.user_id == uid) | (Category.user_id == None),
        Category.cat_type == 'expense'
    ).order_by(Category.position).all()
    cats_income = Category.query.filter(
        (Category.user_id == uid) | (Category.user_id == None),
        Category.cat_type == 'income'
    ).order_by(Category.position).all()
    cards = Card.query.filter_by(user_id=uid).order_by(Card.id).all()

    return render_template('import_categorize.html',
                           rows=data['rows'], skipped_rows=data.get('skipped_rows', []),
                           cats_expense=cats_expense, cats_income=cats_income,
                           cards=cards, tmp=tmp)


@app.route('/import/categorize/confirm', methods=['POST'])
def import_categorize_confirm():
    uid = session.get('user_id')
    if not uid:
        return redirect('/login')
    tmp = request.form.get('tmp', '')
    if not tmp.startswith('impc_'):
        return redirect('/')
    path = os.path.join(tempfile.gettempdir(), tmp)
    if not os.path.exists(path):
        return redirect('/?import_error=세션이 만료되었습니다. 다시 업로드해주세요.')
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    os.remove(path)

    imported = 0
    for i, row in enumerate(data['rows']):
        cat = request.form.get(f'cat_{i}', '기타').strip() or '기타'
        card_val = request.form.get(f'card_{i}', '').strip() or None
        db.session.add(Transaction(
            date=row['date'], type=row['type'], category=cat,
            description=row['description'], amount=row['amount'],
            card=card_val, user_id=uid,
        ))
        imported += 1

    db.session.commit()
    skipped_count = len(data.get('skipped_rows', []))
    return redirect(f'/?imported={imported}&skipped={skipped_count}')

def _parse_sms_line(line):
    amount_m = re.search(r'([\d,]+)원', line)
    if not amount_m:
        return None
    try:
        amount = int(amount_m.group(1).replace(',', ''))
    except Exception:
        return None
    if amount <= 0:
        return None

    d = re.search(r'(\d{4})[-./](\d{1,2})[-./](\d{1,2})', line)
    if d:
        date_val = f"{d.group(1)}-{int(d.group(2)):02d}-{int(d.group(3)):02d}"
    else:
        d = re.search(r'(\d{1,2})[/.-](\d{1,2})', line)
        year = datetime.now().year
        date_val = f"{year}-{int(d.group(1)):02d}-{int(d.group(2)):02d}" if d else datetime.now().strftime('%Y-%m-%d')

    tx_type = 'income' if re.search(r'입금|환급|취소|환불', line) else 'expense'

    bracket_m = re.search(r'\[([^\]]+)\]', line)
    if bracket_m:
        card_val = bracket_m.group(1)
    else:
        card_m = re.search(r'[가-힣a-zA-Z]+(?:카드|은행|뱅크|bank)', line, re.IGNORECASE)
        card_val = card_m.group(0) if card_m else None

    desc = line
    desc = re.sub(r'\[[^\]]+\]', '', desc)
    desc = re.sub(r'[\d,]+원', '', desc)
    desc = re.sub(r'\(금액\)', '', desc)
    desc = re.sub(r'\d{4}[-./]\d{1,2}[-./]\d{1,2}', '', desc)
    desc = re.sub(r'\d{1,2}[/.-]\d{1,2}', '', desc)
    desc = re.sub(r'\d{2}:\d{2}', '', desc)
    desc = re.sub(r'[가-힣]+(?:카드|은행|뱅크)', '', desc)
    desc = re.sub(r'일시불|할부\d*|승인|취소|번호|이체|출금|입금|납부|결제|사용|신용|체크', '', desc)
    desc = re.sub(r'\([^)]*\)', '', desc)
    desc = re.sub(r'[\[\]（）]', '', desc)
    desc = re.sub(r'[가-힣][*]+[가-힣]+', '', desc)
    desc = re.sub(r'\s+', ' ', desc).strip(' -_|,.')

    return {'date': date_val, 'type': tx_type, 'amount': amount,
            'description': desc, 'card': card_val, 'category': '기타'}

@app.route('/import/text', methods=['POST'])
def import_text():
    uid = session.get('user_id')
    if not uid:
        return redirect('/login')
    raw = request.form.get('text', '').strip()
    if not raw:
        return redirect('/?import_error=내용을 입력해주세요')

    parsed = []
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        tx = _parse_sms_line(line)
        if tx:
            parsed.append(tx)

    if not parsed:
        return redirect('/?import_error=인식된 거래 내역이 없습니다')

    fd, path = tempfile.mkstemp(suffix='.json', prefix='impt_')
    with os.fdopen(fd, 'w', encoding='utf-8') as f:
        json.dump(parsed, f, ensure_ascii=False)

    return redirect(url_for('import_text_preview', tmp=os.path.basename(path)))

@app.route('/import/text/preview')
def import_text_preview():
    uid = session.get('user_id')
    if not uid:
        return redirect('/login')
    tmp = request.args.get('tmp', '')
    if not tmp.startswith('impt_'):
        return redirect('/')
    path = os.path.join(tempfile.gettempdir(), tmp)
    if not os.path.exists(path):
        return redirect('/')
    with open(path, encoding='utf-8') as f:
        parsed = json.load(f)
    categories = Category.query.filter_by(user_id=uid).order_by(Category.position, Category.id).all()
    cards = Card.query.filter_by(user_id=uid).all()

    _generic = {'카드', '은행', '뱅크', '체크', '신용', '승인', '출금', '입금', '이체', '결제', '납부'}

    def match_card(sms_card):
        if not sms_card:
            return ''
        for card in cards:
            if card.name in sms_card or sms_card in card.name:
                return card.name
        for card in cards:
            name = card.name
            for length in (2, 3):
                for i in range(len(name) - length + 1):
                    chunk = name[i:i+length]
                    if chunk in _generic:
                        continue
                    if chunk in sms_card:
                        return card.name
        return ''

    for tx in parsed:
        tx['card_matched'] = match_card(tx.get('card', ''))

    expense_cats = [c for c in categories if c.cat_type == 'expense']
    income_cats = [c for c in categories if c.cat_type == 'income']
    return render_template('import_text_preview.html',
                           parsed=parsed, tmp=tmp,
                           categories=categories, card_list=cards,
                           expense_cats_json=[[c.name, c.icon] for c in expense_cats],
                           income_cats_json=[[c.name, c.icon] for c in income_cats])

@app.route('/import/text/confirm', methods=['POST'])
def import_text_confirm():
    uid = session.get('user_id')
    if not uid:
        return redirect('/login')
    tmp = request.form.get('tmp', '')
    if not tmp.startswith('impt_'):
        return redirect('/')
    path = os.path.join(tempfile.gettempdir(), tmp)
    if os.path.exists(path):
        os.remove(path)

    dates  = request.form.getlist('date')
    types  = request.form.getlist('type')
    descs  = request.form.getlist('description')
    amts   = request.form.getlist('amount')
    cats   = request.form.getlist('category')
    cardss = request.form.getlist('card')
    checks = set(request.form.getlist('include'))

    imported = 0
    for i in range(len(dates)):
        if str(i) not in checks:
            continue
        try:
            amount = int(str(amts[i]).replace(',', ''))
            db.session.add(Transaction(
                date=dates[i], type=types[i], category=cats[i],
                description=descs[i], amount=amount,
                card=cardss[i] if cardss[i] else None,
                user_id=uid,
            ))
            imported += 1
        except Exception:
            pass

    db.session.commit()
    return redirect(f'/?imported={imported}')

@app.route('/sw.js')
def service_worker():
    return send_from_directory('static', 'sw.js')

# ── Push Notification ─────────────────────────────────────────────────────────
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_FILE_DIR = os.environ.get('DATA_DIR', _BASE_DIR)
VAPID_PRIVATE_FILE = os.path.join(_FILE_DIR, 'vapid_private_v2.pem')
VAPID_PUBLIC_FILE = os.path.join(_FILE_DIR, 'vapid_public_v2.txt')
SUBSCRIPTIONS_FILE = os.path.join(_FILE_DIR, 'subscriptions.json')

def _get_vapid_keys():
    from cryptography.hazmat.primitives.asymmetric import ec
    from cryptography.hazmat.primitives import serialization
    from cryptography.hazmat.primitives.serialization import load_pem_private_key
    # 기존 키가 PKCS8 형식이면 삭제하고 EC PEM으로 재생성
    if os.path.exists(VAPID_PRIVATE_FILE):
        with open(VAPID_PRIVATE_FILE, 'rb') as f:
            content = f.read()
        if b'BEGIN PRIVATE KEY' in content:  # PKCS8 → 재생성
            for p in [VAPID_PRIVATE_FILE, VAPID_PUBLIC_FILE]:
                if os.path.exists(p): os.remove(p)
    if os.path.exists(VAPID_PRIVATE_FILE) and os.path.exists(VAPID_PUBLIC_FILE):
        try:
            with open(VAPID_PRIVATE_FILE, 'rb') as f:
                load_pem_private_key(f.read(), password=None)
            with open(VAPID_PUBLIC_FILE) as f:
                pub = f.read().strip()
            return {'private': VAPID_PRIVATE_FILE, 'public': pub}
        except Exception:
            for p in [VAPID_PRIVATE_FILE, VAPID_PUBLIC_FILE]:
                if os.path.exists(p): os.remove(p)
    try:
        pk = ec.generate_private_key(ec.SECP256R1())
        # pywebpush가 요구하는 EC PEM 형식 (TraditionalOpenSSL)
        pem = pk.private_bytes(serialization.Encoding.PEM, serialization.PrivateFormat.TraditionalOpenSSL, serialization.NoEncryption())
        with open(VAPID_PRIVATE_FILE, 'wb') as f:
            f.write(pem)
        pub_b64 = base64.urlsafe_b64encode(
            pk.public_key().public_bytes(serialization.Encoding.X962, serialization.PublicFormat.UncompressedPoint)
        ).rstrip(b'=').decode()
        with open(VAPID_PUBLIC_FILE, 'w') as f:
            f.write(pub_b64)
        return {'private': VAPID_PRIVATE_FILE, 'public': pub_b64}
    except Exception:
        return {'private': '', 'public': ''}

vapid_keys = _get_vapid_keys()

def _load_subs():
    if not os.path.exists(SUBSCRIPTIONS_FILE):
        return []
    with open(SUBSCRIPTIONS_FILE) as f:
        return json.load(f)

def _save_subs(subs):
    with open(SUBSCRIPTIONS_FILE, 'w') as f:
        json.dump(subs, f)

@app.route('/api/vapid-public-key')
def vapid_public_key():
    return {'key': vapid_keys['public']}

@app.route('/api/subscribe', methods=['POST'])
def push_subscribe():
    data = request.json or {}
    uid = session.get('user_id')
    if uid:
        data['user_id'] = uid
    subs = [s for s in _load_subs() if s.get('endpoint') != data.get('endpoint')]
    subs.append(data)
    _save_subs(subs)
    return {'ok': True}

@app.route('/api/unsubscribe', methods=['POST'])
def push_unsubscribe():
    data = request.json or {}
    _save_subs([s for s in _load_subs() if s.get('endpoint') != data.get('endpoint')])
    return {'ok': True}

def _do_send_push(sub, title='💰 나의 가계부', body='오늘 지출을 기록했나요? 📝'):
    from pywebpush import webpush
    priv_path = vapid_keys.get('private', '')
    if not priv_path or not os.path.exists(priv_path):
        raise RuntimeError('VAPID private key not found: ' + str(priv_path))
    webpush(
        subscription_info={'endpoint': sub['endpoint'], 'keys': sub['keys']},
        data=json.dumps({'title': title, 'body': body, 'url': '/'}),
        vapid_private_key=priv_path,
        vapid_claims={'sub': 'mailto:song57290@gmail.com'},
    )

def _scheduled_price_update():
    with app.app_context():
        try:
            all_investments = Investment.query.all()
            if all_investments:
                _auto_fetch_investment_prices(all_investments)
        except Exception as e:
            app.logger.error('Scheduled price update error: %s', e)

def _send_push_notifications():
    try:
        from datetime import timezone, timedelta
        KST = timezone(timedelta(hours=9))
        now = datetime.now(KST)
        for sub in _load_subs():
            if now.hour == sub.get('notify_hour', 21) and now.minute == sub.get('notify_minute', 0):
                try:
                    _do_send_push(sub)
                except Exception as e:
                    app.logger.error('Push failed for %s: %s', sub.get('endpoint', '')[:40], e)
    except Exception as e:
        app.logger.error('Push scheduler error: %s', e)

def _send_savings_notifications():
    with app.app_context():
        try:
            from datetime import timezone, timedelta
            KST = timezone(timedelta(hours=9))
            now = datetime.now(KST)
            today_day = now.day
            savings_list = Savings.query.filter_by(stype='청약').all()
            subs = _load_subs()
            for s in savings_list:
                nd = getattr(s, 'notify_day', None)
                if nd and nd == today_day:
                    user_subs = [sub for sub in subs if sub.get('user_id') == s.user_id]
                    for sub in user_subs:
                        try:
                            _do_send_push(
                                sub,
                                title='🏠 청약 납입일 알림',
                                body=f'{s.name} 납입일입니다! {s.amount:,}원을 납입해 주세요.',
                            )
                        except Exception as e:
                            app.logger.error('Savings push failed uid=%s: %s', s.user_id, e)
        except Exception as e:
            app.logger.error('Savings notification error: %s', e)

@app.route('/api/test-notify', methods=['POST'])
@login_required
def test_notify():
    uid = session['user_id']
    subs = [s for s in _load_subs() if s.get('user_id') == uid or True]
    if not subs:
        return jsonify({'ok': False, 'error': '구독 정보 없음. 알림을 먼저 켜주세요.'}), 400
    sent = 0
    errors = []
    for sub in subs:
        try:
            _do_send_push(sub, title='✅ 테스트 알림', body='알림이 정상 작동합니다!')
            sent += 1
        except Exception as e:
            errors.append(str(e))
            app.logger.error('Test push failed: %s', e)
    if sent > 0:
        return jsonify({'ok': True, 'sent': sent})
    return jsonify({'ok': False, 'error': '; '.join(errors)}), 500

if not app.debug or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        import atexit
        _scheduler = BackgroundScheduler()
        _scheduler.add_job(_send_push_notifications, 'cron', minute='*')
        _scheduler.add_job(_send_savings_notifications, 'cron', hour=9, minute=0)
        _scheduler.add_job(_scheduled_price_update, 'cron', day_of_week='mon-fri', hour=6, minute=35)
        _scheduler.add_job(_scheduled_price_update, 'cron', day_of_week='tue-sat', hour=2, minute=15)
        _scheduler.start()
        atexit.register(lambda: _scheduler.shutdown(wait=False))
    except ImportError:
        pass

if __name__ == '__main__':
    app.run(debug=True)
